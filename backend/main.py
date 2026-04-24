"""
main.py — Unified SEO Crawler: FastAPI server + standalone CLI pipeline.
No Streamlit. No external UI dependency.

════════════════════════════════════════════════════════════════════
MODES
════════════════════════════════════════════════════════════════════

1. WEB SERVER (default — serves the HTML dashboard at localhost:8000)
   python main.py
   python main.py --serve --port 8000

2. CLI PIPELINE (crawl → parse → score → AI → export, fully in terminal)
   python main.py --crawl https://example.com
   python main.py --crawl https://example.com --max-pages 30
   python main.py --crawl https://example.com --ai          # run AI after crawl
   python main.py --crawl https://example.com --optimize    # run optimizer
   python main.py --crawl https://example.com --ai --export # export Excel
   python main.py --crawl https://example.com --ai --export --output report.xlsx

3. ENVIRONMENT VARIABLES
   GROQ_API_KEY=gsk_...      (default provider — free at console.groq.com)
   AI_PROVIDER=groq|gemini|openai|claude|ollama|rules
   GROQ_MODEL=llama3-70b-8192

════════════════════════════════════════════════════════════════════
ARCHITECTURE (CLI pipeline)
════════════════════════════════════════════════════════════════════

  crawl()          → async BFS, SSL cascade, extract HTML fields
  parse()          → keywords, scoring, competitor analysis
  generate_prompt()→ build structured Groq prompt per page
  send_to_groq()   → call Groq API (or any provider), return fixes
  display_results()→ print structured table to terminal
  export_excel()   → write full Excel report to disk

════════════════════════════════════════════════════════════════════
FASTAPI ENDPOINTS (web server mode)
════════════════════════════════════════════════════════════════════

  POST /crawl                → start async crawl
  GET  /crawl-status         → live progress
  GET  /results              → full results
  GET  /results/live         → partial results during crawl
  POST /analyze-gemini       → AI on all pages with issues
  POST /analyze-selected     → AI on specific URLs
  GET  /gemini-status        → AI progress
  GET  /gemini-health        → AI provider config
  GET  /ranking/{url}        → instant score
  GET  /popup-data           → pages+fields for popup modal
  POST /optimize             → run live optimization table
  GET  /optimize-status      → optimizer progress
  GET  /optimize-table       → optimization table rows
  GET  /export-optimizer     → download optimization Excel
  GET  /export               → download full report Excel
  GET  /export-popup         → download per-field report Excel
  POST /generate-content     → generate SEO content
  GET  /content-gen-status   → content gen progress
  GET  /generated-content    → all generated content
  GET  /generated-content/{url} → one page's content
  GET  /export-generated-content → download generated content Excel
  GET  /technical-seo            → tech SEO audit for all crawled pages
  GET  /technical-seo/{url}      → tech SEO audit for a single page
  GET  /site-audit               → domain-level health (robots.txt, sitemap, HTTPS)
  GET  /export-technical-seo     → download technical SEO audit Excel

  POST /competitor/analyze        → start competitor analysis task
  GET  /competitor/status/{id}    → poll task status
  GET  /competitor/results/{id}   → full analysis results (when done)
  GET  /competitor/history        → list past analysis snapshots
  DELETE /competitor/{id}         → delete a snapshot
  GET  /competitor/export/{id}    → download competitor report as Excel
"""

# ── Standard library ──────────────────────────────────────────────────────────
import argparse
import asyncio
import io
import json
import logging
import os
import signal
import sys
import tempfile
import time
from collections import Counter
from urllib.parse import unquote, urlparse as _urlparse

# ── Third-party ───────────────────────────────────────────────────────────────
import aiohttp
import pandas as pd
from fastapi import FastAPI, HTTPException, BackgroundTasks, Query, Request as _FastAPIRequest, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from starlette.background import BackgroundTask
import uvicorn
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    _SLOWAPI = True
except ImportError:
    _SLOWAPI = False

# ── Project modules ───────────────────────────────────────────────────────────
from crawler import SEOCrawler, crawl_results, crawl_status
from gemini_analysis import (
    attach_gemini_results, run_gemini_for_pages,
    check_gemini, gemini_status,
    compute_ranking_score,
    _FIELD_RULES, _rule_based_fallback,
    run_content_generation, generate_seo_content,
    content_gen_status, build_seo_content_prompt,
    _rule_based_content, _is_valid_for_content_gen,
)
from seo_optimizer import (
    run_optimization, get_optimization_table,
    clear_optimization_store, optimizer_status,
)
from technical_seo import analyze_page as _tseo_page, analyze_all as _tseo_all
from issues import detect_issues
from keyword_extractor import extract_keywords_corpus
from keyword_scorer import score_keywords, build_structured_page
from competitor_analysis import (
    run_competitor_analysis as _run_comp_analysis,
    get_analysis_result     as _get_comp_result,
    generate_task_id        as _comp_task_id,
)
from competitor_db import list_snapshots as _list_comp_snapshots, delete_snapshot as _del_comp_snapshot

# ── New intelligence modules (graceful fallback if not yet available) ─────────
try:
    from intent_classifier import (
        classify_keywords as _classify_keywords,
        classify_intent   as _classify_intent,
        intent_label      as _intent_label,
    )
    _INTENT_MODULE = True
except ImportError:
    _INTENT_MODULE = False

try:
    from serp_engine import (
        get_ctr_curve                 as _get_ctr_curve,
        expected_ctr                  as _expected_ctr,
        fetch_suggestions_with_intent as _fetch_suggestions_intent,
        score_featured_snippet_potential as _snippet_score,
    )
    _SERP_MODULE = True
except ImportError:
    _SERP_MODULE = False

try:
    from competitor_analysis import detect_cannibalization as _detect_cannibalization
    _CANNIBAL_MODULE = True
except ImportError:
    _CANNIBAL_MODULE = False

try:
    from link_graph import (
        analyse_link_graph as _analyse_link_graph,
        build_link_graph   as _build_link_graph,
        detect_orphans     as _detect_orphans,
    )
    _LINK_GRAPH_MODULE = True
except ImportError:
    _LINK_GRAPH_MODULE = False

try:
    from content_dedup import duplicate_summary as _duplicate_summary
    _DEDUP_MODULE = True
except ImportError:
    _DEDUP_MODULE = False

try:
    from site_auditor import (
        fetch_robots_txt     as _fetch_robots_txt,
        check_hsts           as _check_hsts,
        scan_mixed_content_all as _scan_mixed_content_all,
        run_site_audit       as _run_site_audit,
    )
    _SITE_AUDITOR_MODULE = True
except ImportError:
    _SITE_AUDITOR_MODULE = False

try:
    from serp_scraper import (
        get_serp_position    as _get_serp_position,
        get_keyword_difficulty as _get_keyword_difficulty,
        bulk_serp_check      as _bulk_serp_check,
        bulk_difficulty      as _bulk_difficulty,
    )
    _SERP_SCRAPER_MODULE = True
except ImportError:
    _SERP_SCRAPER_MODULE = False


# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("seo_crawler")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 1: AI PROVIDER HELPERS ──────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def _ai_configured() -> bool:
    """
    Returns True if the currently selected AI provider has its key set.
    Also returns True for 'rules' and 'ollama' which need no key.
    Used by both the FastAPI guards and the CLI pipeline.
    """
    provider = os.getenv("AI_PROVIDER", "groq").lower()
    if provider in ("rules", "ollama"):
        return True
    key_map = {
        "groq":   "GROQ_API_KEY",
        "gemini": "GEMINI_API_KEY",
        "openai": "OPENAI_API_KEY",
        "claude": "ANTHROPIC_API_KEY",
    }
    env_var = key_map.get(provider, "GROQ_API_KEY")
    return bool(os.getenv(env_var, ""))


def _ai_key_error_detail() -> str:
    """Return a human-readable error for the current provider's missing key."""
    provider = os.getenv("AI_PROVIDER", "groq").lower()
    msgs = {
        "groq":   "GROQ_API_KEY not set.  Free key: https://console.groq.com  →  set GROQ_API_KEY=gsk_...",
        "gemini": "GEMINI_API_KEY not set. Run: set GEMINI_API_KEY=your-key",
        "openai": "OPENAI_API_KEY not set. Run: set OPENAI_API_KEY=sk-...",
        "claude": "ANTHROPIC_API_KEY not set. Run: set ANTHROPIC_API_KEY=sk-ant-...",
    }
    return msgs.get(provider, f"API key for provider '{provider}' not set.")


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 2: DISPLAY + FORMATTING HELPERS (formerly in streamlit_app.py) ──
# ════════════════════════════════════════════════════════════════════════════

def priority_icon(p: str) -> str:
    """Return a coloured circle for priority level (terminal-safe)."""
    return {"High": "[HIGH]", "Medium": "[MED]", "Low": "[LOW]"}.get(p, "[---]")


def score_label(s) -> str:
    """Return a letter grade label for a numeric SEO score."""
    if not isinstance(s, (int, float)):
        return "[?]"
    if s >= 70:   return f"[A:{s}]"
    if s >= 45:   return f"[B:{s}]"
    return f"[C:{s}]"


def render_issue_list(issues: list) -> str:
    """Format issues as a comma-separated string for terminal display."""
    return ", ".join(issues) if issues else "None"


def render_kw_importance(kws_scored: list, kws_raw: list) -> str:
    """
    Format keyword list with importance labels for terminal output.
    HIGH keywords appear first and are marked clearly.
    """
    if kws_scored and isinstance(kws_scored[0], dict):
        parts = []
        for s in kws_scored[:8]:
            kw  = s.get("keyword", "")
            imp = s.get("importance", "LOW")
            lbl = {"HIGH": "★", "MEDIUM": "◆", "LOW": "·"}.get(imp, "·")
            parts.append(f"{lbl}{kw}")
        return "  ".join(parts) if parts else "—"
    if kws_raw:
        return "  ".join(f"·{k}" for k in kws_raw[:6])
    return "—"


def display_results(pages: list[dict], show_all: bool = False) -> None:
    """
    Print a structured SEO results table to stdout.
    Replaces the Streamlit expander/table UI.

    Args:
        pages:    list of crawled page dicts
        show_all: if False, only show pages with issues
    """
    pages_to_show = pages if show_all else [p for p in pages if p.get("issues")]

    if not pages_to_show:
        logger.info("No pages to display (no issues found).")
        return

    # Summary counts
    total   = len(pages)
    with_issues = sum(1 for p in pages if p.get("issues"))
    high    = sum(1 for p in pages if p.get("priority") == "High")
    med     = sum(1 for p in pages if p.get("priority") == "Medium")
    clean   = sum(1 for p in pages if not p.get("issues"))

    print()
    print("═" * 80)
    print(f"  SEO CRAWL RESULTS  —  {total} pages crawled")
    print("═" * 80)
    print(f"  With issues:   {with_issues}    High priority: {high}")
    print(f"  Medium:        {med}             Clean:         {clean}")
    print("═" * 80)

    # Issue breakdown
    issue_counts = Counter(i for p in pages for i in p.get("issues", []))
    if issue_counts:
        print()
        print("  ISSUE BREAKDOWN:")
        for issue, count in sorted(issue_counts.items(), key=lambda x: -x[1]):
            bar = "█" * min(count, 30)
            print(f"    {count:3d}×  {issue:<35s}  {bar}")

    print()
    print("─" * 80)
    print(f"  {'PRI':<6} {'SCORE':<8} {'STATUS':<6} {'ISSUES':<35} URL")
    print("─" * 80)

    for page in pages_to_show:
        url      = page.get("url", "")[:65]
        priority = page.get("priority", "")
        issues   = page.get("issues", [])
        ranking  = page.get("ranking") or compute_ranking_score(page)
        score    = ranking.get("score", "?")
        status   = page.get("status_code", "")
        issues_s = render_issue_list(issues)[:33]

        print(f"  {priority_icon(priority):<6} {score_label(score):<8} {str(status):<6} {issues_s:<35} {url}")

        # Show keywords if available
        kws_scored = page.get("keywords_scored") or []
        kws_raw    = page.get("keywords") or []
        if kws_scored or kws_raw:
            kw_str = render_kw_importance(kws_scored, kws_raw)
            print(f"  {'':6} {'':8} {'':6}   KWS: {kw_str[:65]}")

        # Show AI fix preview if available
        gfields = page.get("gemini_fields") or []
        if gfields:
            for f in gfields[:2]:
                if f.get("issue", "OK") != "OK" and f.get("example"):
                    print(f"  {'':6} {'':8} {'':6}   AI→  {f['name']}: {f['example'][:60]}")

        # Show generated content preview
        gc = page.get("generated_content") or {}
        if gc.get("title"):
            print(f"  {'':6} {'':8} {'':6}   GEN: title={gc['title'][:60]}")

    print("─" * 80)
    print()


def display_optimization_table(rows: list[dict]) -> None:
    """
    Print the Live Optimization Table to stdout.
    Replaces the Streamlit dataframe for optimizer results.
    """
    if not rows:
        print("  No optimization rows.")
        return

    print()
    print("═" * 90)
    print("  LIVE OPTIMIZATION TABLE")
    print("═" * 90)
    print(f"  {'FIELD':<20} {'STATUS':<12} {'OPTIMIZED VALUE':<40} URL")
    print("─" * 90)

    for row in rows:
        url   = row.get("url", "")[:40]
        field = row.get("field", "")[:18]
        stat  = row.get("status", "")[:10]
        opt   = row.get("optimized_value", "")[:38]
        print(f"  {field:<20} {stat:<12} {opt:<40} {url}")
        logic = row.get("seo_logic", "")
        if logic:
            print(f"  {'':20} {'':12}   ↳ {logic[:70]}")

    print("─" * 90)
    print()


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 3: EXCEL EXPORT HELPERS ─────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def _style_header(ws) -> None:
    """Apply dark header styling to an openpyxl worksheet."""
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill("solid", fgColor="1A1A2E")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = Font(bold=True, color="00E5A0")
        cell.alignment = Alignment(horizontal="center")


def _style_priority_col(ws) -> None:
    """Colour-code the Priority column in an openpyxl worksheet."""
    from openpyxl.styles import PatternFill, Font
    col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Priority"), None)
    if not col:
        return
    cmap = {"High": "FF4D6A", "Medium": "FFD166", "Low": "06D6A0"}
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        for cell in row:
            if str(cell.value or "") in cmap:
                cell.fill = PatternFill("solid", fgColor=cmap[str(cell.value)])
                cell.font = Font(bold=True, color="000000")


def _style_score_col(ws) -> None:
    """Colour-code the Ranking Score column (green/amber/red)."""
    from openpyxl.styles import PatternFill, Font
    col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Ranking Score"), None)
    if not col:
        return
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        for cell in row:
            try:
                s     = int(cell.value or 0)
                color = "06D6A0" if s >= 70 else "FFD166" if s >= 45 else "FF4D6A"
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="000000")
            except (ValueError, TypeError):
                pass


def _autofit(ws) -> None:
    """Auto-fit column widths in an openpyxl worksheet (max 70 chars)."""
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, 70)


def export_page_ai_to_excel(page: dict) -> bytes | None:
    """
    Build an in-memory Excel file for one page's AI results.
    Includes AI fix fields AND generated content in separate sections.
    Used by the CLI --export flag and the popup export button.
    Returns bytes or None if no AI data exists.
    """
    gc     = page.get("generated_content") or {}
    fields = page.get("gemini_fields") or []
    rows   = []

    # Section 1: AI fix recommendations
    for f in fields:
        rows.append({
            "Section":   "AI Fix",
            "Field":     f.get("name", ""),
            "Issue":     f.get("issue", ""),
            "Current":   f.get("current", ""),
            "Generated": f.get("example") or f.get("fix", ""),
            "Impact":    f.get("impact", ""),
            "Why":       f.get("why", ""),
        })

    # Section 2: Generated content fields
    if gc:
        for fname, fval in [
            ("Title",     gc.get("title", "")),
            ("Meta",      gc.get("meta", "")),
            ("H1",        gc.get("h1", "")),
            ("H2",        " | ".join(gc.get("h2") or [])),
            ("H3",        " | ".join(gc.get("h3") or [])),
            ("Canonical", gc.get("canonical", "")),
            ("Paragraph", gc.get("content", "")),
        ]:
            if fval:
                rows.append({
                    "Section":   "Generated Content",
                    "Field":     fname,
                    "Issue":     "",
                    "Current":   "",
                    "Generated": fval,
                    "Impact":    "",
                    "Why":       gc.get("reason", ""),
                })
        # Section 3: Keyword tracking
        rows.append({
            "Section": "Keywords", "Field": "Used",
            "Generated": ", ".join(gc.get("keywords_used") or []),
            "Issue": "", "Current": "", "Impact": "", "Why": "",
        })
        rows.append({
            "Section": "Keywords", "Field": "Missing",
            "Generated": ", ".join(gc.get("keywords_missing") or []),
            "Issue": "", "Current": "", "Impact": "", "Why": "",
        })

    if not rows:
        return None

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="AI Results")
    buf.seek(0)
    return buf.getvalue()


def export_excel(pages: list[dict], output_path: str) -> None:
    """
    Export a full SEO report to an Excel file at output_path.
    Called from the CLI --export flag.
    Includes all pages, issues, keywords, scores, and AI fixes.
    """
    rows = []
    for r in pages:
        ranking = r.get("ranking") or compute_ranking_score(r)
        # Collect any AI-generated fixes for this page
        ai_fix = "; ".join(
            f"{f['name']}: {f.get('fix', '')}"
            for f in (r.get("gemini_fields") or [])
            if f.get("fix") and f.get("issue", "OK") != "OK"
        )
        rows.append({
            "url":              r.get("url", ""),
            "status_code":      r.get("status_code", ""),
            "title":            r.get("title", ""),
            "meta_description": r.get("meta_description", ""),
            "h1":               " | ".join(r.get("h1") or []),
            "h2":               " | ".join((r.get("h2") or [])[:3]),
            "canonical":        r.get("canonical", ""),
            "keywords":         ", ".join(r.get("keywords") or []),
            "competition":      r.get("competition", ""),
            "internal_links":   r.get("internal_links_count", 0),
            "issues":           ", ".join(r.get("issues") or []),
            "priority":         r.get("priority", ""),
            "ranking_score":    ranking.get("score", ""),
            "ranking_grade":    ranking.get("grade", ""),
            "ai_fix":           ai_fix,
            "generated_title":   (r.get("generated_content") or {}).get("title", ""),
            "generated_content": (r.get("generated_content") or {}).get("content", "")[:500],
            "faq_count":         len((r.get("generated_content") or {}).get("faq") or []),
        })

    if not rows:
        logger.warning("No crawl data to export.")
        return

    df = pd.DataFrame(rows)
    df.columns = [c.replace("_", " ").title() for c in df.columns]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SEO Report")
        ws = writer.sheets["SEO Report"]
        _style_header(ws)
        _style_priority_col(ws)
        _style_score_col(ws)
        _autofit(ws)

    logger.info("Report exported: %s  (%d rows)", output_path, len(rows))


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 4: GROQ PROMPT GENERATION ───────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def generate_prompt(page: dict) -> str:
    """
    Build a structured Groq-ready SEO prompt for one page.

    Uses ONLY data extracted from the crawled page — never hallucinated.
    Includes: URL, title, meta, h1/h2/h3, keywords with importance scores,
    competitor gaps, detected issues, and first 1000 chars of body text.

    Returns a formatted string ready to send to send_to_groq().
    """
    return build_seo_content_prompt(page)


def send_to_groq(page: dict) -> dict:
    """
    Send a structured SEO prompt to the configured AI provider and return results.

    Provider is selected via AI_PROVIDER env var (default: groq).
    Falls back to rule-based generation if API is unavailable.

    Returns a dict with keys:
        url, title, meta, h1, h2, h3, canonical, content,
        keywords_used, keywords_missing, reason, _source
    """
    if not _is_valid_for_content_gen(page):
        logger.debug("Page skipped (no content/keywords): %s", page.get("url"))
        return _rule_based_content(page)

    if not _ai_configured():
        logger.warning("AI not configured — using rule-based fallback.")
        return _rule_based_content(page)

    try:
        result = generate_seo_content(page)
        provider = os.getenv("AI_PROVIDER", "groq").upper()
        source   = result.get("_source", "unknown")
        logger.info("AI result (%s→%s): %s", provider, source, page.get("url", "")[:60])
        return result
    except Exception as exc:
        logger.error("send_to_groq failed for %s: %s", page.get("url"), exc)
        return _rule_based_content(page)


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 5: CLI PIPELINE ──────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

async def _crawl_async(url: str, max_pages: int) -> list[dict]:
    """
    Internal async BFS crawl. Handles SSL cascade, redirects, errors.
    Updates crawl_results and crawl_status in-place via the crawler module.
    Returns snapshot of crawl_results after completion.
    """
    crawl_results.clear()
    clear_optimization_store()

    # Normalise URL scheme
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    crawl_status.update({
        "running": True, "done": False,
        "pages_crawled": 0, "pages_queued": 0,
        "errors": 0, "timeouts": 0, "ssl_fallbacks": 0,
        "current_url": "", "error": None,
        "started_at": time.time(), "elapsed_s": 0,
    })

    try:
        await SEOCrawler(url, max_pages=max_pages).crawl_async()
    except Exception as exc:
        logger.error("Crawl error: %s", exc)
        crawl_status.update({"running": False, "done": False, "error": str(exc)})

    return list(crawl_results)


def crawl(url: str, max_pages: int = 50) -> list[dict]:
    """
    Public synchronous crawl function for CLI use.

    Runs the async BFS crawler in a new event loop.
    Handles HTTP errors, SSL failures, redirects, and timeouts gracefully.
    Returns a list of page dicts (both successful and error pages).

    Each page dict contains:
        url, status_code, title, meta_description, h1, h2, h3,
        canonical, body_text, internal_links_count,
        keywords, keywords_scored, competitor_gaps, issues, priority,
        competition, _is_error, gemini_fields, generated_content
    """
    logger.info("Starting crawl: %s  (max %d pages)", url, max_pages)
    t0 = time.time()

    # Run async crawler in a fresh event loop
    pages = asyncio.run(_crawl_async(url, max_pages))

    elapsed = round(time.time() - t0, 1)
    real    = [p for p in pages if p.get("status_code") == 200]
    errors  = [p for p in pages if p.get("_is_error")]

    logger.info(
        "Crawl complete: %d pages in %ss  (%d real, %d errors, %d SSL fallbacks)",
        len(pages), elapsed, len(real), len(errors),
        crawl_status.get("ssl_fallbacks", 0),
    )
    return pages


def parse(pages: list[dict]) -> list[dict]:
    """
    Post-crawl parsing pipeline — runs on results from crawl().

    Steps:
      1. detect_issues()          — find SEO problems on every page
      2. extract_keywords_corpus() — TF-IDF across all real pages
      3. score_keywords()         — HIGH/MEDIUM/LOW importance per keyword
      4. build_structured_page()  — canonical output shape

    Modifies pages in-place (adds issues, keywords, keywords_scored, structured).
    Returns the same list with all fields populated.
    """
    real = [p for p in pages if not p.get("_is_error") and p.get("status_code") == 200]
    logger.info("Parsing %d real pages (of %d total)…", len(real), len(pages))

    # Step 1: detect SEO issues
    detect_issues(pages)

    # Step 2: TF-IDF keyword extraction across full corpus
    if real:
        extract_keywords_corpus(real, top_n=10)

    # Step 3: score keywords and build structured output per page
    for page in real:
        try:
            scored = score_keywords(page, top_n=10)
            page["keywords_scored"] = scored
            page["structured"]      = build_structured_page(page, scored_keywords=scored)
        except Exception as exc:
            logger.warning("parse() failed for %s: %s", page.get("url"), exc)
            page.setdefault("keywords_scored", [])
            page.setdefault("structured", None)

    # Assign priority to all pages (including errors)
    from gemini_analysis import assign_priority
    for page in pages:
        page["priority"] = assign_priority(page.get("issues", []))

    logger.info(
        "Parse complete: %d keywords extracted, %d pages with issues",
        sum(len(p.get("keywords", [])) for p in real),
        sum(1 for p in pages if p.get("issues")),
    )
    return pages


def run_ai_pipeline(pages: list[dict], max_pages: int = 10) -> list[dict]:
    """
    Run AI analysis on pages with issues.

    Uses the configured AI provider (default: Groq).
    Falls back to rule-based if no API key is set.
    Processes up to max_pages to stay within free-tier rate limits.

    Args:
        pages:     crawl results from crawl() + parse()
        max_pages: maximum pages to send to AI (default 10)

    Returns the same list with gemini_fields populated on each page.
    """
    provider = os.getenv("AI_PROVIDER", "groq").upper()
    logger.info("Running AI pipeline (%s) on up to %d pages…", provider, max_pages)

    valid = [
        p for p in pages
        if not p.get("_is_error")
        and p.get("status_code") == 200
        and p.get("issues")
        and (p.get("body_text") or "")
    ][:max_pages]

    if not valid:
        logger.info("No valid pages for AI analysis.")
        return pages

    if not _ai_configured():
        logger.warning("AI not configured (%s). Skipping AI pipeline.", _ai_key_error_detail())
        return pages

    # Run attach_gemini_results synchronously (it uses ThreadPoolExecutor internally)
    try:
        attach_gemini_results(valid)
        # Merge results back into the main pages list
        url_map = {p["url"]: p for p in pages}
        for page in valid:
            if page.get("gemini_fields") and page["url"] in url_map:
                url_map[page["url"]]["gemini_fields"] = page["gemini_fields"]
                url_map[page["url"]]["ranking"]       = compute_ranking_score(page)
        logger.info("AI pipeline complete: %d pages analysed", len(valid))
    except Exception as exc:
        logger.error("AI pipeline failed: %s", exc)

    return pages


def run_optimizer_pipeline(pages: list[dict]) -> list[dict]:
    """
    Run the Live Optimization Table generator on pages with issues.

    Generates paste-ready optimized values (title, meta, H1, etc.) for
    every broken field. Falls back to rule-based if AI unavailable.

    Returns optimizer rows (not attached to pages — use get_optimization_table()).
    """
    logger.info("Running optimizer pipeline…")
    try:
        run_optimization(pages)
        rows = get_optimization_table()
        logger.info("Optimizer complete: %d rows generated", len(rows))
        return rows
    except Exception as exc:
        logger.error("Optimizer pipeline failed: %s", exc)
        return []


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 6: CLI ENTRY POINT ───────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def _run_cli(args: argparse.Namespace) -> None:
    """
    Standalone CLI pipeline: crawl → parse → AI → display → export.
    No server, no browser needed. Runs entirely in the terminal.
    """
    # ── Step 1: crawl() ───────────────────────────────────────────────────
    pages = crawl(args.crawl, max_pages=args.max_pages)
    if not pages:
        logger.error("Crawl returned no results. Check the URL and network.")
        sys.exit(1)

    # ── Step 2: parse() ───────────────────────────────────────────────────
    pages = parse(pages)

    # ── Step 3: AI analysis (optional, --ai flag) ─────────────────────────
    if args.ai:
        pages = run_ai_pipeline(pages, max_pages=args.max_ai_pages)

    # ── Step 4: optimizer (optional, --optimize flag) ─────────────────────
    opt_rows = []
    if args.optimize:
        opt_rows = run_optimizer_pipeline(pages)

    # ── Step 5: display_results() ─────────────────────────────────────────
    display_results(pages, show_all=args.show_all)
    if opt_rows:
        display_optimization_table(opt_rows)

    # ── Step 6: export Excel (optional, --export flag) ────────────────────
    if args.export:
        export_excel(pages, args.output)

        # Also export generated content if AI ran
        if args.ai:
            gen_rows = []
            for p in pages:
                gc = p.get("generated_content")
                if gc:
                    gen_rows.append({
                        "URL":     gc.get("url", ""),
                        "Title":   gc.get("title", ""),
                        "Meta":    gc.get("meta", ""),
                        "H1":      gc.get("h1", ""),
                        "Content": gc.get("content", ""),
                        "Source":  gc.get("_source", ""),
                        "Reason":  gc.get("reason", ""),
                    })
            if gen_rows:
                gen_path = args.output.replace(".xlsx", "_generated.xlsx")
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    pd.DataFrame(gen_rows).to_excel(writer, index=False,
                                                    sheet_name="Generated Content")
                buf.seek(0)
                with open(gen_path, "wb") as f:
                    f.write(buf.read())
                logger.info("Generated content exported: %s", gen_path)

    # ── Final summary ──────────────────────────────────────────────────────
    total      = len(pages)
    issues_ct  = sum(1 for p in pages if p.get("issues"))
    high_ct    = sum(1 for p in pages if p.get("priority") == "High")
    ai_ct      = sum(1 for p in pages if p.get("gemini_fields"))

    print()
    print("═" * 60)
    print("  DONE")
    print(f"  Pages crawled:     {total}")
    print(f"  Pages with issues: {issues_ct}  ({high_ct} high priority)")
    if args.ai:
        print(f"  AI analysed:       {ai_ct}")
    if args.export:
        print(f"  Report saved:      {args.output}")
    print("═" * 60)
    print()


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 7: FASTAPI SERVER (web mode) ─────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

app = FastAPI(title="SEO Crawler API")

# ── Rate limiter (brute force protection on auth endpoints) ──────────────────
# slowapi is optional — if not installed the app still starts, auth just has
# no per-IP throttle (acceptable for local dev, required for production).
if _SLOWAPI:
    _limiter = Limiter(key_func=get_remote_address)
    app.state.limiter = _limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    logger.info("Rate limiter active (slowapi): auth endpoints limited to 5/minute per IP.")

    def _rate_limit(limit_str: str):
        """Apply slowapi rate limit decorator."""
        return _limiter.limit(limit_str)
else:
    def _rate_limit(limit_str: str):  # type: ignore[misc]
        """No-op when slowapi is not installed — passes through unchanged."""
        def _noop(func):
            return func
        return _noop

# BUG-006 / BUG-N16: restrict CORS — read from env so production can lock it down.
# Default stays open for local dev / Hugging Face Space usage.
_allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")
if "*" in _allowed_origins:
    logger.warning(
        "CORS is open to all origins (ALLOWED_ORIGINS=*). "
        "Set ALLOWED_ORIGINS=https://yourdomain.com before exposing this server publicly."
    )
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-API-Key"],
)

# BUG-001: lock prevents two simultaneous /crawl requests corrupting shared state.
_crawl_lock = asyncio.Lock()


def _delete_tempfile(path: str) -> None:
    """BUG-004: delete temp export file after response is fully sent."""
    try:
        os.unlink(path)
    except OSError:
        pass


# ── Health check (BUG-019: k8s readiness/liveness probe) ────────────────────

@app.get("/healthz")
def health_check():
    """Kubernetes readiness/liveness probe. Always returns 200 when the process is up."""
    return {
        "status":        "ok",
        "crawl_running": crawl_status.get("running", False),
        "pages_crawled": len(crawl_results),
        "ai_provider":   os.getenv("AI_PROVIDER", "groq"),
        "ai_configured": _ai_configured(),
    }


# ── Static assets ─────────────────────────────────────────────────────────────
# Mount /static and /backend/static to the same dir so both relative-path
# conventions work: tool pages use "../static/js/…" which resolves to either
# /static/… (when served at /pages/) or /backend/static/… (when served at
# /backend/pages/) — both point to the same physical directory.
_static_dir = os.path.join(BASE_DIR, "static")
if os.path.isdir(_static_dir):
    app.mount("/static", StaticFiles(directory=_static_dir), name="static")
    app.mount("/backend/static", StaticFiles(directory=_static_dir), name="backend_static")

def _read_html(path: str) -> HTMLResponse:
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())

# ── Frontend: landing page ────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def serve_ui():
    """Serve the new landing page; fall back to legacy dashboard if absent."""
    landing = os.path.join(BASE_DIR, "landing.html")
    legacy  = os.path.join(BASE_DIR, "index.html")
    return _read_html(landing if os.path.exists(landing) else legacy)

# ── Frontend: tool pages ──────────────────────────────────────────────────────
# Serve at BOTH /pages/<name> and /backend/pages/<name> so the same relative
# links work whether users land on GitHub Pages or directly on the HF Space.
def _serve_page(page_name: str) -> HTMLResponse:
    if not page_name.endswith(".html"):
        raise HTTPException(status_code=404, detail="Not found")
    page_path = os.path.join(BASE_DIR, "pages", page_name)
    if not os.path.isfile(page_path):
        raise HTTPException(status_code=404, detail="Page not found")
    return _read_html(page_path)

@app.get("/pages/{page_name}", response_class=HTMLResponse)
def serve_page(page_name: str):
    return _serve_page(page_name)

@app.get("/backend/pages/{page_name}", response_class=HTMLResponse)
def serve_backend_page(page_name: str):
    return _serve_page(page_name)

# ── Legacy dashboard (kept for direct /dashboard access) ─────────────────────
@app.get("/dashboard", response_class=HTMLResponse)
def serve_dashboard():
    return _read_html(os.path.join(BASE_DIR, "index.html"))


# ── Request models ────────────────────────────────────────────────────────────

class CrawlRequest(BaseModel):
    # BUG-N04: min/max length rejects empty strings and 1MB+ URLs before
    # any processing happens.
    url:       str = Field(..., min_length=10, max_length=2048)
    # BUG-002: enforce 1–500 range to prevent DoS via huge crawl requests.
    max_pages: int = Field(50, ge=1, le=500)

# BUG-N04: private / cloud-metadata hostnames the crawler must never reach.
_SSRF_BLOCKED = {
    "localhost", "127.0.0.1", "0.0.0.0",
    "169.254.169.254",          # AWS / GCP / Azure instance metadata
    "metadata.google.internal", # GCP metadata alias
}

class SelectedPagesRequest(BaseModel):
    urls: list[str]


# ── Crawl endpoints ───────────────────────────────────────────────────────────

@app.post("/crawl")
async def start_crawl(request: CrawlRequest, http_request: _FastAPIRequest = None):
    """
    Fire-and-forget async crawl.
    Returns immediately — poll /crawl-status every 2s for progress.
    """
    url = request.url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    # BUG-N04: SSRF guard — reject private/metadata IP addresses and hostnames.
    _parsed_url = _urlparse(url)
    _host = (_parsed_url.hostname or "").lower()
    if _host in _SSRF_BLOCKED or _host.startswith(("192.168.", "10.", "172.16.")):
        raise HTTPException(
            status_code=400,
            detail="Private or reserved IP addresses are not allowed.",
        )

    # Quota check for authenticated users
    _crawl_user_id = None
    if _AUTH_MODULE and http_request is not None:
        _crawl_user = _get_current_user(http_request)
        if _crawl_user:
            _crawl_user_id = _crawl_user["id"]
            _allowed, _msg = _auth_check_quota(_crawl_user_id, request.max_pages)
            if not _allowed:
                raise HTTPException(status_code=429, detail=_msg)

    # BUG-N07 + BUG-001: the entire state reset is now inside the lock so no
    # concurrent request ever sees a window where running=True but results are
    # from the previous crawl.
    async with _crawl_lock:
        if crawl_status.get("running"):
            raise HTTPException(status_code=409, detail="Crawl already running.")
        # Reset all state under the lock — prevents readers seeing stale data
        crawl_results.clear()
        clear_optimization_store()
        crawl_status.update({
            "running": True, "done": False,
            "pages_crawled": 0, "pages_queued": 0,
            "errors": 0, "timeouts": 0, "ssl_fallbacks": 0,
            "current_url": "", "error": None,
            "started_at": None, "elapsed_s": 0,
        })
        content_gen_status.update({
            "running": False, "done": False, "error": None,
            "processed": 0, "total": 0,
        })
        gemini_status.update({
            "running": False, "done": False, "error": None,
            "processed": 0, "total": 0, "skipped": 0,
        })

    asyncio.create_task(_run_crawl(url, request.max_pages, _crawl_user_id))
    return {"status": "running", "message": "Crawl started"}


async def _run_crawl(url: str, max_pages: int, user_id: int | None = None) -> None:
    """Background task: runs the async crawler, logs errors, records quota usage."""
    try:
        await SEOCrawler(url, max_pages=max_pages).crawl_async()
        # Assign priority from issues detected by crawl_async() → detect_issues().
        # This must run here because the CLI parse() path is not called via the API.
        from gemini_analysis import assign_priority
        for page in crawl_results:
            if not page.get("priority"):
                page["priority"] = assign_priority(page.get("issues", []))
        # Record pages crawled against the user's monthly quota
        if user_id and _AUTH_MODULE:
            pages_done = len(crawl_results)
            if pages_done > 0:
                _auth_record_pages(user_id, pages_done)
    except Exception as exc:
        logger.error("Crawl failed: %s", exc)
        crawl_status.update({"running": False, "done": False, "error": str(exc)})


@app.get("/crawl-status")
def get_crawl_status():
    """Live crawl progress — poll every 2s from frontend."""
    return crawl_status


@app.get("/results")
def get_results(
    limit:  int = Query(default=0, ge=0, description="Max pages to return (0 = all)"),
    offset: int = Query(default=0, ge=0, description="Number of pages to skip"),
):
    """
    Full results after crawl completes.
    BUG-007: supports optional pagination via limit/offset query params
    (limit=0 means return all — preserves existing behaviour for the frontend).
    BUG-N26: limit/offset now validated by Query(ge=0) — negative values return 422.
    """
    all_results = list(crawl_results)
    sliced = all_results[offset: offset + limit] if limit > 0 else all_results[offset:]
    return {
        "status":        crawl_status,
        "gemini_status": gemini_status,
        "total":         len(all_results),
        "results":       sliced,
    }


@app.get("/results/live")
def get_results_live():
    """Pages crawled so far — safe to call mid-crawl for live table updates."""
    # BUG-N18: use "total" (not "count") — same key as /results for consistency.
    snapshot = list(crawl_results)
    return {
        "status":  crawl_status,
        "total":   len(snapshot),
        "results": snapshot,
    }


# ── AI endpoints ──────────────────────────────────────────────────────────────

@app.post("/analyze-gemini")
async def analyze_gemini():
    """Run AI (Groq/Gemini/etc) on top-20 pages with issues. Non-blocking."""
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data yet.")
    if gemini_status.get("running"):
        raise HTTPException(status_code=409, detail="AI already running.")
    if not _ai_configured():
        raise HTTPException(status_code=400, detail=_ai_key_error_detail())

    loop = asyncio.get_running_loop()
    asyncio.create_task(_run_gemini_executor(loop, list(crawl_results)))
    return {"message": "AI analysis started."}


@app.post("/analyze-selected")
async def analyze_selected(request: SelectedPagesRequest):
    """Run AI on user-selected URLs only."""
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data yet.")
    if gemini_status.get("running"):
        raise HTTPException(status_code=409, detail="AI already running.")
    if not request.urls:
        raise HTTPException(status_code=400, detail="No URLs provided.")
    if not _ai_configured():
        raise HTTPException(status_code=400, detail=_ai_key_error_detail())

    loop = asyncio.get_running_loop()
    asyncio.create_task(
        _run_selected_executor(loop, request.urls, list(crawl_results))
    )
    return {"message": f"AI analysis started for {len(request.urls)} pages."}


async def _run_gemini_executor(loop, snapshot: list[dict]) -> None:
    """Thread pool wrapper so AI SDK calls don't block the event loop."""
    try:
        await loop.run_in_executor(None, attach_gemini_results, snapshot)
        _merge_gemini_results(snapshot)
    except Exception as exc:
        gemini_status.update({"error": str(exc), "running": False, "done": False})


async def _run_selected_executor(loop, urls: list[str], snapshot: list[dict]) -> None:
    """Thread pool wrapper for per-URL AI analysis."""
    try:
        await loop.run_in_executor(None, run_gemini_for_pages, urls, snapshot)
        _merge_gemini_results(snapshot)
    except Exception as exc:
        gemini_status.update({"error": str(exc), "running": False, "done": False})


def _merge_gemini_results(snapshot: list[dict]) -> None:
    """
    Write AI results from the snapshot back into live crawl_results.
    Called after AI analysis completes in the thread pool.

    BUG-N23: build the url_map from a snapshot copy so we never iterate
    crawl_results directly while another coroutine may be clearing it.
    We write back using index lookup on the same snapshot-derived map,
    which is safe because list entries are mutable dicts shared by reference.
    """
    # Snapshot the list reference once — safe even if crawl_results is replaced
    current = list(crawl_results)
    url_map = {p["url"]: p for p in current}
    for page in snapshot:
        url = page["url"]
        if url in url_map:
            if page.get("gemini_fields"):
                url_map[url]["gemini_fields"] = page["gemini_fields"]
            if page.get("priority"):
                url_map[url]["priority"] = page["priority"]
            url_map[url]["ranking"] = compute_ranking_score(url_map[url])


@app.get("/gemini-status")
def get_gemini_status():
    """Poll AI analysis progress."""
    return gemini_status


@app.get("/gemini-health")
def gemini_health():
    """AI provider health check — provider-aware, works for all backends."""
    info     = check_gemini()
    provider = os.getenv("AI_PROVIDER", "groq").lower()
    info["provider"]   = provider
    info["configured"] = _ai_configured()
    if provider == "groq":
        info["model"]    = os.getenv("GROQ_MODEL", "llama3-70b-8192")
        key              = os.getenv("GROQ_API_KEY", "")
        info["key_hint"] = ("..." + key[-4:]) if len(key) > 8 else "(not set)"
    return info


# ── AI Key Setup endpoints ────────────────────────────────────────────────────

# Map of provider → env variable that holds its key
_PROVIDER_KEY_MAP = {
    "groq":   "GROQ_API_KEY",
    "gemini": "GEMINI_API_KEY",
    "openai": "OPENAI_API_KEY",
    "claude": "ANTHROPIC_API_KEY",
    "ollama": None,   # no key needed — local
    "rules":  None,   # no key needed — rule-based only
}

_PROVIDER_LABELS = {
    "groq":   "Groq (Llama 3)",
    "gemini": "Google Gemini",
    "openai": "OpenAI GPT-4o-mini",
    "claude": "Anthropic Claude",
    "ollama": "Ollama (Local)",
    "rules":  "Rule-based (no AI)",
}


class SetApiKeyRequest(BaseModel):
    provider: str
    api_key: str = ""


@app.get("/ai-config")
def get_ai_config():
    """
    Return current AI provider configuration.
    Used by the frontend AI Setup popup to show current state.
    """
    provider = os.getenv("AI_PROVIDER", "gemini").lower()
    env_var  = _PROVIDER_KEY_MAP.get(provider)
    key      = os.getenv(env_var, "") if env_var else ""
    key_hint = ("..." + key[-4:]) if len(key) > 8 else ("(set)" if key else "(not set)")
    configured = bool(key) or provider in ("ollama", "rules")
    return {
        "provider":   provider,
        "label":      _PROVIDER_LABELS.get(provider, provider),
        "configured": configured,
        "key_hint":   key_hint,
        "providers":  list(_PROVIDER_LABELS.items()),   # [{provider, label}, ...]
    }


@app.post("/set-api-key")
def set_api_key(req: SetApiKeyRequest):
    """
    Set the AI provider and its API key at runtime.

    Updates:
      - os.environ so new key lookups via os.getenv() see the change immediately
      - gemini_analysis.AI_PROVIDER and seo_optimizer.AI_PROVIDER module-level
        variables so in-flight provider routing switches without a server restart

    The key is never logged or returned — only the last-4 hint is echoed back.
    """
    provider = req.provider.lower().strip()
    if provider not in _PROVIDER_KEY_MAP:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown provider '{provider}'. Valid: {', '.join(_PROVIDER_KEY_MAP)}"
        )

    # Update env vars
    os.environ["AI_PROVIDER"] = provider
    env_var = _PROVIDER_KEY_MAP.get(provider)
    api_key = req.api_key.strip()
    if api_key and env_var:
        os.environ[env_var] = api_key

    # Update module-level AI_PROVIDER in both AI modules so the change
    # takes effect without a server restart (modules are singletons in Python).
    import gemini_analysis as _ga
    import seo_optimizer   as _so
    _ga.AI_PROVIDER = provider
    _so.AI_PROVIDER = provider

    key_hint = ("..." + api_key[-4:]) if len(api_key) > 8 else ("(set)" if api_key else "(not set)")
    configured = bool(api_key) or provider in ("ollama", "rules")

    logger.info("AI provider updated to '%s' via /set-api-key", provider)
    return {
        "ok":         True,
        "provider":   provider,
        "label":      _PROVIDER_LABELS.get(provider, provider),
        "configured": configured,
        "key_hint":   key_hint,
    }


# ── Ranking + popup ───────────────────────────────────────────────────────────

@app.get("/ranking/{page_url:path}")
def get_ranking(page_url: str):
    """Instant SEO score for a single page (no API call)."""
    decoded = unquote(page_url)
    page    = next((p for p in crawl_results if p["url"] == decoded), None)
    if not page:
        raise HTTPException(status_code=404, detail="Page not found.")
    return compute_ranking_score(page)


@app.get("/popup-data")
def get_popup_data():
    """Return pages-with-issues formatted for the fix-page popup modal."""
    popup_pages = []
    for p in crawl_results:
        if not p.get("issues"):
            continue
        ranking = p.get("ranking") or compute_ranking_score(p)
        if p.get("gemini_ranking_score") is not None:
            ranking = dict(ranking)
            ranking["gemini_score"]  = p["gemini_ranking_score"]
            ranking["gemini_reason"] = p.get("gemini_ranking_reason", "")
        popup_pages.append({
            "url":         p["url"],
            "priority":    p.get("priority", ""),
            "issues":      p.get("issues", []),
            "keywords":    p.get("keywords", []),
            "competition": p.get("competition", "Medium"),
            "ranking":     ranking,
            "fields":      _build_popup_fields(p),
        })
    return {"total": len(popup_pages), "pages": popup_pages}


def _build_popup_fields(page: dict) -> list[dict]:
    """
    Build per-field rows for the popup modal.
    Prefers AI-generated fixes; falls back to rule-based suggestions.
    """
    issues_set = set(page.get("issues", []))
    gemini_map = {f["name"]: f for f in (page.get("gemini_fields") or [])}
    opt_title  = page.get("optimized_title", "")
    opt_meta   = page.get("optimized_meta", "")
    opt_h1     = page.get("optimized_h1", "")

    fields_def = [
        ("Title",            page.get("title", ""),            _title_status(issues_set)),
        ("Meta Description", page.get("meta_description", ""), _meta_status(issues_set)),
        ("H1",               (page.get("h1") or [""])[0],      _h1_status(issues_set)),
        ("H2",               " | ".join((page.get("h2") or [])[:2]),
                             "Missing" if "Missing H2" in issues_set else "OK"),
        ("Canonical",        page.get("canonical", ""),        _canonical_status(issues_set)),
        ("URL",              page.get("url", ""),               "OK"),
    ]

    rows = []
    for name, current, status in fields_def:
        gf        = gemini_map.get(name)
        opt_value = {"Title": opt_title, "Meta Description": opt_meta, "H1": opt_h1}.get(name, "")

        if gf and gf.get("fix"):
            row = {
                "field": name, "current": current or "", "status": status,
                "why":   gf.get("why", ""), "fix": gf.get("fix", ""),
                "example": gf.get("example", ""), "impact": gf.get("impact", ""),
                "optimized": opt_value or gf.get("example", ""),
            }
        elif status != "OK":
            rule = _FIELD_RULES.get(name, {}).get(status, {})
            row = {
                "field": name, "current": current or "", "status": status,
                "why":   rule.get("why", ""), "fix": rule.get("fix", f"Review {name}."),
                "example": rule.get("example", ""), "impact": rule.get("impact", ""),
                "optimized": opt_value,
            }
        else:
            row = {
                "field": name, "current": current or "", "status": "OK",
                "why": "", "fix": "", "example": "", "impact": "", "optimized": "",
            }
        rows.append(row)
    return rows


# ── Status field helpers ──────────────────────────────────────────────────────

def _title_status(s: set) -> str:
    # BUG-N38: "Title Too Short" was falling through to "OK" — popup showed
    # wrong status, suppressing the fix-suggestion row for short titles.
    if "Missing Title"   in s: return "Missing"
    if "Title Too Long"  in s: return "Too Long"
    if "Title Too Short" in s: return "Too Short"
    return "OK"

def _meta_status(s: set) -> str:
    # BUG-N44: "Meta Description Too Long" and "Too Short" were both mapped
    # to "OK" — popup suppressed fix rows for length issues.
    if "Missing Meta Description"    in s: return "Missing"
    if "Duplicate Meta Description"  in s: return "Duplicate"
    if "Meta Description Too Long"   in s: return "Too Long"
    if "Meta Description Too Short"  in s: return "Too Short"
    return "OK"

def _h1_status(s: set) -> str:
    if "Missing H1"       in s: return "Missing"
    if "Multiple H1 Tags" in s: return "Multiple"
    return "OK"

def _canonical_status(s: set) -> str:
    if "Missing Canonical"  in s: return "Missing"
    if "Canonical Mismatch" in s: return "Mismatch"
    return "OK"


# ── Optimizer endpoints ───────────────────────────────────────────────────────

class OptimizeRequest(BaseModel):
    urls: list[str] | None = None


@app.post("/optimize")
async def start_optimize(request: OptimizeRequest):
    """Run seo_optimizer on crawled pages. Non-blocking."""
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data yet.")
    if optimizer_status.get("running"):
        raise HTTPException(status_code=409, detail="Optimizer already running.")
    if not _ai_configured():
        raise HTTPException(status_code=400, detail=_ai_key_error_detail())

    # BUG-N31: clear stale rows before each run so re-running /optimize
    # on the same crawl session never accumulates rows from prior runs.
    clear_optimization_store()

    snapshot = list(crawl_results)
    urls     = request.urls if request.urls else None
    loop     = asyncio.get_running_loop()
    asyncio.create_task(_run_optimizer_executor(loop, snapshot, urls))
    return {"message": "Optimizer started.", "pages": len(snapshot)}


async def _run_optimizer_executor(loop, snapshot, urls) -> None:
    """Thread pool wrapper for the optimizer."""
    try:
        await loop.run_in_executor(None, run_optimization, snapshot, urls)
    except Exception as exc:
        optimizer_status.update({"error": str(exc), "running": False, "done": False})


@app.get("/optimize-status")
def get_optimize_status():
    """Poll optimizer progress."""
    return optimizer_status


@app.get("/optimize-table")
def get_optimize_table():
    """Return the full Live Optimization Table."""
    from seo_optimizer import _optimization_store
    return {
        "status": optimizer_status,
        "total":  sum(len(rows) for rows in _optimization_store.values()),
        "rows":   get_optimization_table(),
    }


# ── Excel export endpoints ────────────────────────────────────────────────────

@app.get("/export-optimizer")
def export_optimizer_excel():
    """Download the Live Optimization Table as Excel."""
    rows = get_optimization_table()
    if not rows:
        raise HTTPException(status_code=404, detail="No optimization data. Run /optimize first.")

    df        = pd.DataFrame(rows)
    col_order = ["url", "field", "status", "current_value", "optimized_value", "seo_logic"]
    df        = df[[c for c in col_order if c in df.columns]]
    df.columns = [c.replace("_", " ").title() for c in df.columns]

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Live Optimization")
        ws = writer.sheets["Live Optimization"]
        _style_header(ws)
        status_col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Status"), None)
        if status_col:
            from openpyxl.styles import PatternFill, Font
            for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
                for cell in row:
                    v = str(cell.value or "").lower()
                    if v in ("missing", "too long", "duplicate", "multiple", "mismatch"):
                        cell.fill = PatternFill("solid", fgColor="FF4D6A")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif v == "ok":
                        cell.fill = PatternFill("solid", fgColor="06D6A0")
                        cell.font = Font(color="000000", bold=True)
        _autofit(ws)

    # BUG-004: delete temp file after the response is fully streamed.
    return FileResponse(tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="seo_optimization_table.xlsx",
        headers={"Content-Disposition": "attachment; filename=seo_optimization_table.xlsx"},
        background=BackgroundTask(_delete_tempfile, tmp.name))


@app.get("/export")
def export_excel_endpoint():
    """Download full SEO report as Excel."""
    if not crawl_results:
        raise HTTPException(status_code=404, detail="No crawl data to export.")

    rows = []
    for r in crawl_results:
        ranking = r.get("ranking") or compute_ranking_score(r)
        ai_fix  = "; ".join(
            f"{f['name']}: {f.get('fix', '')}"
            for f in (r.get("gemini_fields") or [])
            if f.get("fix") and f.get("issue", "OK") != "OK"
        )
        rows.append({
            "url":              r.get("url", ""),
            "status_code":      r.get("status_code", ""),
            "title":            r.get("title", ""),
            "meta_description": r.get("meta_description", ""),
            "h1":               " | ".join(r.get("h1") or []),
            "h2":               " | ".join((r.get("h2") or [])[:3]),
            "canonical":        r.get("canonical", ""),
            "keywords":         ", ".join(r.get("keywords") or []),
            "competition":      r.get("competition", ""),
            "internal_links":   r.get("internal_links_count", 0),
            "issues":           ", ".join(r.get("issues") or []),
            "priority":         r.get("priority", ""),
            "ranking_score":    ranking["score"],
            "ranking_grade":    ranking["grade"],
            "ai_fix":           ai_fix,
        })

    df  = pd.DataFrame(rows)
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SEO Report")
        ws = writer.sheets["SEO Report"]
        _style_header(ws)
        _style_priority_col(ws)
        _style_score_col(ws)
        _autofit(ws)

    # BUG-004: delete temp file after the response is fully streamed.
    return FileResponse(tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="seo_report.xlsx",
        headers={"Content-Disposition": "attachment; filename=seo_report.xlsx"},
        background=BackgroundTask(_delete_tempfile, tmp.name))


@app.get("/export-popup")
def export_popup_excel():
    """Download per-field issues report as Excel."""
    popup = get_popup_data()
    if not popup["pages"]:
        raise HTTPException(status_code=404, detail="No pages with issues to export.")

    rows = []
    for page in popup["pages"]:
        score = page.get("ranking", {}).get("score", "")
        for f in page["fields"]:
            rows.append({
                "URL":             page["url"],
                "Priority":        page.get("priority", ""),
                "Score":           score,
                "Field":           f["field"],
                "Current Value":   f["current"],
                "Issue":           f["status"],
                "Why It Matters":  f.get("why", ""),
                "Exact Fix":       f.get("fix", ""),
                "Optimized Value": f.get("optimized", ""),
                "Real Example":    f.get("example", ""),
                "Impact":          f.get("impact", ""),
            })

    df  = pd.DataFrame(rows)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SEO Issues")
        ws = writer.sheets["SEO Issues"]
        _style_header(ws)
        issue_col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Issue"), None)
        if issue_col:
            from openpyxl.styles import PatternFill, Font
            for row in ws.iter_rows(min_row=2, min_col=issue_col, max_col=issue_col):
                for cell in row:
                    val = str(cell.value or "").upper()
                    if val == "OK":
                        cell.fill = PatternFill("solid", fgColor="06D6A0")
                        cell.font = Font(color="000000", bold=True)
                    elif val:
                        cell.fill = PatternFill("solid", fgColor="FF4D6A")
                        cell.font = Font(color="FFFFFF", bold=True)
        _autofit(ws)

    # BUG-004: delete temp file after the response is fully streamed.
    return FileResponse(tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="seo_field_report.xlsx",
        headers={"Content-Disposition": "attachment; filename=seo_field_report.xlsx"},
        background=BackgroundTask(_delete_tempfile, tmp.name))


# ── Content generation endpoints ──────────────────────────────────────────────

@app.post("/generate-content")
async def start_content_generation(background_tasks: BackgroundTasks):
    """Generate complete SEO content for all crawled pages. Non-blocking."""
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl results. Run a crawl first.")

    def _run():
        run_content_generation(crawl_results)

    background_tasks.add_task(_run)
    valid_count = sum(
        1 for p in crawl_results
        if not p.get("_is_error") and p.get("status_code") == 200
        and (p.get("keywords_scored") or p.get("keywords"))
    )
    return {"message": "Content generation started", "pages_queued": valid_count}


@app.get("/content-gen-status")
def get_content_gen_status():
    """Poll content generation progress."""
    return content_gen_status


@app.get("/generated-content")
def get_generated_content():
    """Return all generated SEO content."""
    pages_with_content = [
        {"url": p.get("url", ""), "issues": p.get("issues", []),
         "generated": p.get("generated_content")}
        for p in crawl_results if p.get("generated_content") is not None
    ]
    return {"total": len(pages_with_content), "status": content_gen_status,
            "pages": pages_with_content}


@app.get("/generated-content/{page_url:path}")
def get_page_generated_content(page_url: str):
    """Get generated content for a single page by URL."""
    page = next((p for p in crawl_results if p.get("url") == page_url), None)
    if not page:
        raise HTTPException(status_code=404, detail="Page not found")
    gc = page.get("generated_content")
    if not gc:
        raise HTTPException(status_code=404, detail="No generated content for this page yet")
    return gc


@app.get("/export-generated-content")
def export_generated_content():
    """Export all generated content as Excel."""
    rows = []
    for p in crawl_results:
        gc = p.get("generated_content")
        if not gc:
            continue
        rows.append({
            "URL":               gc.get("url", ""),
            "Generated Title":   gc.get("title", ""),
            "Generated Meta":    gc.get("meta", ""),
            "Generated H1":      gc.get("h1", ""),
            "Generated H2s":     " | ".join(gc.get("h2") or []),
            "Generated H3s":     " | ".join(gc.get("h3") or []),
            "Canonical":         gc.get("canonical", ""),
            "Generated Content": gc.get("content", ""),
            "Keywords Used":     ", ".join(gc.get("keywords_used") or []),
            "Keywords Missing":  ", ".join(gc.get("keywords_missing") or []),
            "Reason":            gc.get("reason", ""),
            "Source":            gc.get("_source", ""),
        })

    if not rows:
        raise HTTPException(status_code=404, detail="No generated content to export")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Generated Content")
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=generated_seo_content.xlsx"},
    )


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 7b: TECHNICAL SEO ENDPOINTS ─────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

@app.get("/technical-seo")
def get_technical_seo():
    """Return technical SEO audit for all crawled pages (post-crawl only)."""
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data yet.")
    return _tseo_all(list(crawl_results))


@app.get("/technical-seo/{page_url:path}")
def get_technical_seo_page(page_url: str):
    """Return technical SEO audit for a single crawled page.
    BUG-011: unquote() normalises percent-encoded URLs so paths containing
    %2F, %3F, %23 etc. match the stored URL string correctly.
    """
    decoded = unquote(page_url)
    page = next((p for p in crawl_results if p.get("url") == decoded), None)
    if not page:
        raise HTTPException(status_code=404, detail="URL not found in crawl results.")
    return _tseo_page(page)


@app.get("/site-audit")
async def get_site_audit():
    """
    Fetch domain-level technical health signals for the crawled site.

    Makes at most 2 async HTTP requests (robots.txt + sitemap.xml) using a
    5-second timeout each. Returns:
      domain        — origin URL
      robots_txt    — {status, accessible, blocks_googlebot, content_preview}
      sitemap       — {status, accessible, url}
      https_summary — derived from crawl_results (no extra requests)
    """
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data yet.")

    first_url = crawl_results[0].get("url", "")
    parsed    = _urlparse(first_url)
    domain    = f"{parsed.scheme}://{parsed.netloc}"

    robots_url  = f"{domain}/robots.txt"
    sitemap_url = f"{domain}/sitemap.xml"

    timeout = aiohttp.ClientTimeout(total=5)
    headers = {"User-Agent": "CrawlIQ-TechSEO/1.0 (+https://crawliq.io)"}

    async def _fetch(url: str) -> tuple[int, str]:
        """Return (status_code, body_text). Body capped at 4000 chars."""
        try:
            async with aiohttp.ClientSession(timeout=timeout, headers=headers) as sess:
                async with sess.get(url, ssl=False, allow_redirects=True) as resp:
                    body = await resp.text(errors="replace")
                    return resp.status, body[:4000]
        except Exception as exc:
            return 0, str(exc)

    robots_status, robots_body = await _fetch(robots_url)
    sitemap_status, sitemap_body = await _fetch(sitemap_url)

    # ── Robots.txt analysis ───────────────────────────────────────────────────
    robots_accessible = robots_status == 200
    blocks_googlebot  = False
    disallow_all      = False
    if robots_accessible:
        # BUG-N20: accumulate agents across consecutive User-agent lines so a
        # multi-agent block like:
        #   User-agent: Googlebot
        #   User-agent: Bingbot
        #   Disallow: /
        # correctly detects that Googlebot is blocked (not just Bingbot).
        lines = [ln.strip() for ln in robots_body.splitlines()]
        current_agents: list[str] = []
        _prev_was_directive = False
        for ln in lines:
            low = ln.lower()
            if low.startswith("user-agent:"):
                agent = ln.split(":", 1)[1].strip().lower()
                # A blank line or directive before this line starts a new group
                if _prev_was_directive:
                    current_agents = []
                    _prev_was_directive = False
                current_agents.append(agent)
            elif low.startswith(("disallow:", "allow:")):
                path = ln.split(":", 1)[1].strip()
                if path == "/" and any(a in ("*", "googlebot") for a in current_agents):
                    blocks_googlebot = True
                    disallow_all     = True
                _prev_was_directive = True
            elif ln == "":
                # Blank line ends the current agent group
                current_agents = []
                _prev_was_directive = False

    robots_result = {
        "url":               robots_url,
        "accessible":        robots_accessible,
        "status_code":       robots_status,
        "blocks_googlebot":  blocks_googlebot,
        "disallow_all":      disallow_all,
        "content_preview":   robots_body[:800] if robots_accessible else "",
        "status":            (
            "blocks_crawlers" if blocks_googlebot
            else "ok" if robots_accessible
            else "not_found" if robots_status == 404
            else "error"
        ),
    }

    # ── Sitemap analysis ──────────────────────────────────────────────────────
    sitemap_accessible = sitemap_status == 200
    sitemap_is_xml     = sitemap_accessible and (
        "<?xml" in sitemap_body or "<urlset" in sitemap_body or "<sitemapindex" in sitemap_body
    )
    # Rough URL count: count <loc> tags
    url_count = sitemap_body.count("<loc>") if sitemap_is_xml else 0

    sitemap_result = {
        "url":         sitemap_url,
        "accessible":  sitemap_accessible,
        "status_code": sitemap_status,
        "is_xml":      sitemap_is_xml,
        "url_count":   url_count,
        "status":      (
            "ok" if sitemap_is_xml
            else "not_xml" if sitemap_accessible
            else "not_found" if sitemap_status == 404
            else "error"
        ),
    }

    # ── HTTPS summary from crawl data (no extra requests) ────────────────────
    real_pages   = [p for p in crawl_results if not p.get("_is_error")]
    https_pages  = sum(1 for p in real_pages if (p.get("url") or "").startswith("https://"))
    total_real   = len(real_pages) or 1
    https_pct    = round((https_pages / total_real) * 100, 1)

    https_summary = {
        "total_pages":  len(crawl_results),
        "real_pages":   len(real_pages),
        "https_pages":  https_pages,
        "http_pages":   len(real_pages) - https_pages,
        "https_pct":    https_pct,
        "status":       (
            "all_https"  if https_pct == 100
            else "partial" if https_pct > 0
            else "no_https"
        ),
    }

    # ── HTTP status code distribution from crawl data ────────────────────────
    status_dist: dict[str, int] = {}
    for p in crawl_results:
        code = str(p.get("status_code", "Unknown"))
        bucket = (
            "2xx" if code.startswith("2") else
            "3xx" if code.startswith("3") else
            "4xx" if code.startswith("4") else
            "5xx" if code.startswith("5") else
            code
        )
        status_dist[bucket] = status_dist.get(bucket, 0) + 1

    return {
        "domain":         domain,
        "robots_txt":     robots_result,
        "sitemap":        sitemap_result,
        "https_summary":  https_summary,
        "status_distribution": status_dist,
    }


@app.get("/export-technical-seo")
def export_technical_seo():
    """
    Export the full technical SEO audit as a styled Excel file.
    Sheet 1: Per-page technical audit (score, grade, indexability, per-component scores)
    Sheet 2: Site-wide summary
    """
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data yet.")

    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    audit = _tseo_all(list(crawl_results))
    pages = audit["pages"]
    summary = audit["summary"]

    # ── Sheet 1: Per-page rows ────────────────────────────────────────────────
    page_rows = []
    for a in pages:
        # BUG-N27: use .get() throughout so a partial audit dict never raises KeyError.
        idx      = a.get("indexability", {})
        title_a  = a.get("title", {})
        meta_a   = a.get("meta", {})
        canon_a  = a.get("canonical", {})
        head_a   = a.get("headings", {})
        og_a     = a.get("open_graph", {})
        cont_a   = a.get("content", {})
        url_a    = a.get("url_analysis", {})
        img_a    = a.get("images", {})
        page_rows.append({
            "URL":              a.get("url", ""),
            "HTTP Status":      a.get("status_code", ""),
            "Indexability":     idx.get("label", "Unknown"),
            "Index Reason":     idx.get("reason", ""),
            "Tech Score":       a.get("tech_score", 0),
            "Tech Grade":       a.get("tech_grade", "?"),
            "Issue Count":      a.get("issue_count", 0),
            "All Issues":       " | ".join(a.get("all_issues", [])),
            # Component scores
            "Title Score":      title_a.get("score", 0),
            "Title Status":     title_a.get("status", ""),
            "Title Value":      title_a.get("value", ""),
            "Meta Score":       meta_a.get("score", 0),
            "Meta Status":      meta_a.get("status", ""),
            "Meta Value":       meta_a.get("value", ""),
            "Canonical Score":  canon_a.get("score", 0),
            "Canonical Status": canon_a.get("status", ""),
            "H1 Count":         head_a.get("h1_count", 0),
            "H2 Count":         head_a.get("h2_count", 0),
            "Heading Score":    head_a.get("score", 0),
            "OG Score":         og_a.get("score", 0),
            "OG Status":        og_a.get("completeness", ""),
            "Content Words":    cont_a.get("word_count", 0),
            "Content Depth":    cont_a.get("depth", ""),
            "Content Score":    cont_a.get("score", 0),
            "URL HTTPS":        url_a.get("is_https", False),
            "URL Depth":        url_a.get("depth", 0),
            "URL Score":        url_a.get("score", 0),
            "Image Score":      img_a.get("score", 0),
            "Image Status":     img_a.get("status", ""),
        })

    df_pages = pd.DataFrame(page_rows)

    # ── Sheet 2: Summary rows ─────────────────────────────────────────────────
    cov   = summary.get("coverage", {})
    idx_s = summary.get("indexability", {})
    cont  = summary.get("content", {})

    summary_rows = [
        {"Metric": "Total Pages",           "Value": summary.get("total_pages", 0)},
        {"Metric": "Real Pages",             "Value": summary.get("real_pages", 0)},
        {"Metric": "Avg Tech Score",         "Value": summary.get("avg_tech_score", 0)},
        {"Metric": "Site Grade",             "Value": summary.get("site_grade", "")},
        {"Metric": "Indexable Pages",        "Value": idx_s.get("indexable_total", 0)},
        {"Metric": "Indexable %",            "Value": idx_s.get("indexable_pct", 0)},
        {"Metric": "Blocked / Error Pages",  "Value": idx_s.get("blocked_total", 0)},
        {"Metric": "Canonical Mismatch",     "Value": idx_s.get("canonical_mismatch", 0)},
        {"Metric": "Title Coverage %",       "Value": cov.get("title_pct", 0)},
        {"Metric": "Meta Coverage %",        "Value": cov.get("meta_pct", 0)},
        {"Metric": "Canonical Coverage %",   "Value": cov.get("canonical_pct", 0)},
        {"Metric": "OG Coverage %",          "Value": cov.get("og_pct", 0)},
        {"Metric": "H1 Coverage %",          "Value": cov.get("h1_pct", 0)},
        {"Metric": "HTTPS Coverage %",       "Value": cov.get("https_pct", 0)},
        {"Metric": "Thin Content Pages",     "Value": cont.get("thin_pages", 0)},
        {"Metric": "Thin Content %",         "Value": cont.get("thin_pct", 0)},
    ]
    for item in summary.get("top_issues", []):
        summary_rows.append({"Metric": f"Issue: {item['issue']}", "Value": item["count"]})

    df_summary = pd.DataFrame(summary_rows)

    # ── Build Excel ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_pages.to_excel(writer, index=False, sheet_name="Technical SEO")
        df_summary.to_excel(writer, index=False, sheet_name="Site Summary")

        # Style Sheet 1
        ws1 = writer.sheets["Technical SEO"]
        hdr_fill = PatternFill("solid", fgColor="0D1B2A")
        for cell in ws1[1]:
            cell.fill      = hdr_fill
            cell.font      = Font(bold=True, color="22D3EE")
            cell.alignment = Alignment(horizontal="center")

        # Colour-code Tech Score column
        score_col = next(
            (i for i, c in enumerate(ws1[1], 1) if c.value == "Tech Score"), None
        )
        if score_col:
            for row in ws1.iter_rows(min_row=2, min_col=score_col, max_col=score_col):
                for cell in row:
                    try:
                        s = int(cell.value or 0)
                        fgColor = "06D6A0" if s >= 70 else "FFD166" if s >= 40 else "FF4D6A"
                        cell.fill = PatternFill("solid", fgColor=fgColor)
                        cell.font = Font(bold=True, color="000000")
                    except (ValueError, TypeError):
                        pass

        # Colour-code Indexability column
        idx_col = next(
            (i for i, c in enumerate(ws1[1], 1) if c.value == "Indexability"), None
        )
        if idx_col:
            idx_colors = {
                "Indexable": ("C8F7C5", "000000"),
                "Likely":    ("D4EFDF", "000000"),
                "Redirect":  ("FEF9E7", "000000"),
                "Canonical": ("FDEBD0", "000000"),
            }
            for row in ws1.iter_rows(min_row=2, min_col=idx_col, max_col=idx_col):
                for cell in row:
                    v = str(cell.value or "")
                    for key, (bg, fg) in idx_colors.items():
                        if key in v:
                            cell.fill = PatternFill("solid", fgColor=bg)
                            cell.font = Font(bold=True, color=fg)
                            break

        # Autofit Sheet 1
        for col in ws1.columns:
            w = max(len(str(c.value or "")) for c in col) + 4
            ws1.column_dimensions[get_column_letter(col[0].column)].width = min(w, 60)

        # Style Sheet 2
        ws2 = writer.sheets["Site Summary"]
        for cell in ws2[1]:
            cell.fill      = PatternFill("solid", fgColor="0D1B2A")
            cell.font      = Font(bold=True, color="22D3EE")
            cell.alignment = Alignment(horizontal="center")
        for col in ws2.columns:
            w = max(len(str(c.value or "")) for c in col) + 4
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(w, 50)

    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=technical_seo_audit.xlsx"},
    )


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 7c: COMPETITOR ANALYSIS ENDPOINTS ────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

class CompetitorRequest(BaseModel):
    target_url:       str = Field(..., min_length=10, max_length=2048)
    competitor_urls:  list[str] = Field(..., min_length=1, max_length=5)


@app.post("/competitor/analyze")
async def start_competitor_analysis(request: CompetitorRequest):
    """
    Start a competitor analysis task.
    Crawls target + competitor sites, fetches Core Web Vitals from PSI,
    scores 7 dimensions, computes keyword gaps and radar chart data.

    Returns immediately with a task_id — poll /competitor/status/{task_id}.
    Analysis typically completes in 30–90 seconds.
    """
    target = request.target_url.strip()
    if not target.startswith(("http://", "https://")):
        target = "https://" + target

    # SSRF guard
    _host = (_urlparse(target).hostname or "").lower()
    if _host in _SSRF_BLOCKED or _host.startswith(("192.168.", "10.", "172.16.")):
        raise HTTPException(status_code=400, detail="Private IPs not allowed.")

    # Normalise and validate competitor URLs
    competitors = []
    for raw in request.competitor_urls:
        u = raw.strip()
        if not u:
            continue
        if not u.startswith(("http://", "https://")):
            u = "https://" + u
        parsed_comp = _urlparse(u)
        if not parsed_comp.netloc or "." not in parsed_comp.netloc:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid competitor URL (no valid domain): {raw!r}",
            )
        comp_host = parsed_comp.hostname or ""
        if comp_host in _SSRF_BLOCKED or comp_host.startswith(("192.168.", "10.", "172.16.")):
            raise HTTPException(status_code=400, detail=f"Private/blocked host not allowed: {raw!r}")
        competitors.append(u)

    if not competitors:
        raise HTTPException(status_code=400, detail="At least one competitor URL required.")
    if len(competitors) > 5:
        raise HTTPException(status_code=400, detail="Maximum 5 competitor URLs.")

    task_id = _comp_task_id()

    # Persist pending row before starting background task
    from competitor_db import save_snapshot as _save_snap
    _save_snap(task_id, target, competitors)

    # Run analysis as background task (non-blocking)
    asyncio.create_task(_run_comp_analysis(task_id, target, competitors))

    logger.info("Competitor analysis started: task=%s target=%s competitors=%d",
                task_id, target, len(competitors))
    return {
        "task_id":          task_id,
        "status":           "running",
        "target_url":       target,
        "competitor_urls":  competitors,
        "message":          "Analysis started. Poll /competitor/status/" + task_id,
    }


@app.get("/competitor/status/{task_id}")
def get_competitor_status(task_id: str):
    """
    Poll competitor analysis task status.
    Returns {task_id, status, created_at, completed_at, error_msg}.
    status: pending | running | done | error
    """
    snap = _get_comp_result(task_id)
    if snap is None:
        raise HTTPException(status_code=404, detail=f"Task '{task_id}' not found.")
    return {
        "task_id":      snap["task_id"],
        "status":       snap["status"],
        "created_at":   snap.get("created_at"),
        "completed_at": snap.get("completed_at"),
        "error_msg":    snap.get("error_msg"),
    }


@app.get("/competitor/results/{task_id}")
def get_competitor_results(task_id: str):
    """
    Return full competitor analysis results once status == 'done'.
    Raises 404 if not found, 202 if still running.
    """
    snap = _get_comp_result(task_id)
    if snap is None:
        raise HTTPException(status_code=404, detail=f"Task '{task_id}' not found.")
    if snap["status"] == "running":
        import json as _json
        return Response(
            content=_json.dumps({
                "status":  "running",
                "message": f"Analysis still in progress. Poll /competitor/status/{task_id}",
            }),
            status_code=202,
            media_type="application/json",
        )
    if snap["status"] == "error":
        # BUG-009: log full error server-side; never return raw internal state to client.
        logger.error("Competitor task %s failed: %s", task_id, snap.get("error_msg", ""))
        raise HTTPException(status_code=500, detail="Analysis failed — see server logs.")
    return {
        "task_id":   snap["task_id"],
        "status":    snap["status"],
        "created_at": snap.get("created_at"),
        "completed_at": snap.get("completed_at"),
        "results":   snap.get("metrics", {}),
    }


@app.get("/competitor/history")
def get_competitor_history(
    domain: str | None = Query(default=None, description="Filter by domain substring"),
    limit:  int        = Query(default=20, ge=1, le=100),
):
    """
    List past competitor analysis snapshots, newest first.
    Optionally filter by target domain substring.
    """
    snapshots = _list_comp_snapshots(domain=domain, limit=limit)
    return {
        "total":     len(snapshots),
        "snapshots": [
            {
                "task_id":      s["task_id"],
                "target_url":   s["target_url"],
                "competitor_urls": s.get("competitor_urls", []),
                "status":       s["status"],
                "created_at":   s.get("created_at"),
                "completed_at": s.get("completed_at"),
                "summary":      s.get("summary", {}),
            }
            for s in snapshots
        ],
    }


@app.delete("/competitor/{task_id}")
def delete_competitor_snapshot(task_id: str):
    """Hard-delete a competitor analysis snapshot and all related data."""
    deleted = _del_comp_snapshot(task_id)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"Task '{task_id}' not found.")
    return {"ok": True, "task_id": task_id}


@app.get("/competitor/export/{task_id}")
def export_competitor_excel(task_id: str):
    """
    Download competitor analysis as a styled multi-sheet Excel file.
    Sheets: Summary Scores | Keyword Gaps | E-E-A-T | Core Web Vitals | Actions
    """
    snap = _get_comp_result(task_id)
    if snap is None:
        raise HTTPException(status_code=404, detail="Task not found.")
    if snap["status"] != "done":
        raise HTTPException(status_code=400, detail="Analysis not yet complete.")

    metrics       = snap.get("metrics") or {}
    sites         = metrics.get("sites", [])
    if not sites:
        raise HTTPException(
            status_code=422,
            detail="Analysis metrics are incomplete — no site data to export. "
                   "The analysis may have failed silently; please re-run it.",
        )
    gaps          = metrics.get("keyword_gaps", [])
    actions       = metrics.get("actions", [])
    target        = metrics.get("target_url", "")
    crawl_errors  = metrics.get("crawl_errors", {})

    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Score Comparison ─────────────────────────────────────
        score_rows = []
        for site in sites:
            sc      = site.get("scores", {})
            blocked = site.get("crawl_blocked", False)
            na      = "N/A (crawl blocked)" if blocked else None

            source = site.get("score_source", "crawl")

            def _s(key):
                if source == "no_data":
                    return "N/A (no data)"
                return sc.get(key, 0)

            score_rows.append({
                "Domain":        site.get("domain", ""),
                "URL":           site.get("url", ""),
                "Composite":     _s("composite"),
                "Technical":     _s("technical"),
                "On-Page":       _s("on_page"),
                "Content":       _s("content"),
                "E-E-A-T":       _s("eeat"),
                "CTR Potential": _s("ctr"),
                "Keywords":      _s("keywords"),
                "Page Speed":    _s("page_speed"),
                "Is Target":     "✓" if site.get("url") == target else "",
                "Score Source":  {"crawl": "Crawled", "psi_derived": "PSI/Lighthouse", "no_data": "No Data"}.get(source, source),
                "Pages Crawled": site.get("real_pages", 0) if source == "crawl" else "—",
            })
        df_scores = pd.DataFrame(score_rows)
        df_scores.to_excel(writer, index=False, sheet_name="Score Comparison")
        ws = writer.sheets["Score Comparison"]
        _style_header(ws)
        _autofit(ws)
        # Colour Composite column
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                try:
                    s = float(cell.value or 0)
                    fgColor = "06D6A0" if s >= 70 else "FFD166" if s >= 45 else "FF4D6A"
                    cell.fill = PatternFill("solid", fgColor=fgColor)
                    cell.font = Font(bold=True, color="000000")
                except (ValueError, TypeError):
                    pass

        # ── Sheet 2: Keyword Gaps ─────────────────────────────────────────
        if gaps:
            gap_rows = [
                {
                    "Keyword":            g["keyword"],
                    "Competitor Count":   g["competitor_count"],
                    "Found In":           ", ".join(g.get("found_in", [])),
                    "Opportunity Score":  g["opportunity_score"],
                }
                for g in gaps[:100]
            ]
            pd.DataFrame(gap_rows).to_excel(writer, index=False, sheet_name="Keyword Gaps")
            ws2 = writer.sheets["Keyword Gaps"]
            _style_header(ws2)
            _autofit(ws2)

        # ── Sheet 3: Core Web Vitals ──────────────────────────────────────
        cwv_rows = []
        for site in sites:
            cwv = site.get("cwv", {})
            if cwv:
                cwv_rows.append({
                    "Domain":       site.get("domain", ""),
                    "Perf Score":   cwv.get("perf_score", ""),
                    "LCP (ms)":     cwv.get("lcp_ms", ""),
                    "LCP Status":   cwv.get("lcp_status", ""),
                    "CLS":          cwv.get("cls", ""),
                    "CLS Status":   cwv.get("cls_status", ""),
                    "FCP (ms)":     cwv.get("fcp_ms", ""),
                    "TTFB (ms)":    cwv.get("ttfb_ms", ""),
                    "Field Data":   cwv.get("field_data", ""),
                    "Strategy":     cwv.get("strategy", "mobile"),
                })
        if cwv_rows:
            pd.DataFrame(cwv_rows).to_excel(writer, index=False, sheet_name="Core Web Vitals")
            ws3 = writer.sheets["Core Web Vitals"]
            _style_header(ws3)
            _autofit(ws3)

        # ── Sheet 4: Action Priority List ─────────────────────────────────
        if actions:
            act_rows = [
                {
                    "Priority":             a["priority"],
                    "Dimension":            a["label"],
                    "Target Score":         a["target_score"],
                    "Avg Competitor Score": a["avg_competitor_score"],
                    "Gap":                  a["gap"],
                    "Recommended Action":   a["action"],
                }
                for a in actions
            ]
            pd.DataFrame(act_rows).to_excel(writer, index=False, sheet_name="Action Plan")
            ws4 = writer.sheets["Action Plan"]
            _style_header(ws4)
            _autofit(ws4)

        # ── Sheet 5: Crawl Errors (only if any sites were blocked) ────────
        if crawl_errors:
            err_rows = [
                {
                    "URL":    url,
                    "Domain": _urlparse(url).netloc,
                    "Reason": reason,
                    "Fix":    "Add delay/User-Agent rotation, or use Playwright for JS-heavy sites",
                }
                for url, reason in crawl_errors.items()
            ]
            pd.DataFrame(err_rows).to_excel(writer, index=False, sheet_name="Crawl Errors")
            ws5 = writer.sheets["Crawl Errors"]
            _style_header(ws5)
            _autofit(ws5)

    buf.seek(0)
    safe_domain = (_urlparse(target).netloc or "competitor").replace(".", "_")
    filename    = f"competitor_analysis_{safe_domain}.xlsx"

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(buf.getvalue())
    tmp.close()

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
        background=BackgroundTask(_delete_tempfile, tmp.name),
    )


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 8: ENTRYPOINT ────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def _build_arg_parser() -> argparse.ArgumentParser:
    """Build the CLI argument parser for both server and pipeline modes."""
    parser = argparse.ArgumentParser(
        prog="main.py",
        description=(
            "SEO Crawler Dashboard — FastAPI server + standalone CLI pipeline.\n"
            "No Streamlit required.\n\n"
            "Examples:\n"
            "  python main.py                              # start web server\n"
            "  python main.py --crawl https://example.com # CLI crawl\n"
            "  python main.py --crawl https://example.com --ai --export\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    # Server mode
    server_group = parser.add_argument_group("Server mode (default)")
    server_group.add_argument(
        "--serve", action="store_true",
        help="Explicitly start the FastAPI server (default when no --crawl given)",
    )
    server_group.add_argument(
        "--port", type=int, default=8000,
        help="Port for the web server (default: 8000)",
    )
    server_group.add_argument(
        "--host", default="0.0.0.0",
        help="Host for the web server (default: 0.0.0.0)",
    )
    server_group.add_argument(
        "--reload", action="store_true",
        help="Enable hot-reload during development",
    )

    # CLI pipeline mode
    cli_group = parser.add_argument_group("CLI pipeline mode")
    cli_group.add_argument(
        "--crawl", metavar="URL",
        help="URL to crawl (activates CLI pipeline mode)",
    )
    cli_group.add_argument(
        "--max-pages", type=int, default=50, metavar="N",
        help="Maximum pages to crawl (default: 50)",
    )
    cli_group.add_argument(
        "--ai", action="store_true",
        help="Run AI analysis after crawl (requires API key)",
    )
    cli_group.add_argument(
        "--max-ai-pages", type=int, default=10, metavar="N",
        help="Maximum pages to send to AI per run (default: 10, free tier safe)",
    )
    cli_group.add_argument(
        "--optimize", action="store_true",
        help="Run optimizer after crawl to generate paste-ready fixes",
    )
    cli_group.add_argument(
        "--export", action="store_true",
        help="Export results to Excel after pipeline completes",
    )
    cli_group.add_argument(
        "--output", default="seo_report.xlsx", metavar="FILE",
        help="Output Excel filename (default: seo_report.xlsx)",
    )
    cli_group.add_argument(
        "--show-all", action="store_true",
        help="Show all pages in terminal, not just pages with issues",
    )

    return parser


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION N: INTELLIGENCE ENDPOINTS (intent / SERP / cannibalization) ──────
# ════════════════════════════════════════════════════════════════════════════

@app.get("/serp/ctr-curve")
def get_ctr_curve_endpoint():
    """
    Return Sistrix 2024 CTR benchmark curve for positions 1-20.

    Response:
      {
        "curve": [
          {"position": 1, "ctr_pct": 28.5, "ctr_frac": 0.285,
           "delta_vs_prev": null, "tier": "gold", "tier_color": "#f59e0b"},
          ...
        ]
      }

    Use this to render a CTR opportunity chart in the frontend.
    """
    if not _SERP_MODULE:
        raise HTTPException(status_code=503, detail="serp_engine module not available.")
    return {"curve": _get_ctr_curve()}


@app.get("/serp/suggestions")
async def get_serp_suggestions(q: str = "", lang: str = "en"):
    """
    Fetch keyword suggestions from Google Suggest with intent classification.

    Query params:
      q    — seed keyword (required)
      lang — language code, default "en"

    Response:
      {
        "keyword":     "seo tools",
        "suggestions": [
          {"suggestion": "best seo tools 2026", "intent": "commercial",
           "intent_short": "Comm", "intent_color": "#f59e0b"},
          ...
        ]
      }
    """
    if not q.strip():
        raise HTTPException(status_code=400, detail="Query parameter 'q' is required.")
    if not _SERP_MODULE:
        raise HTTPException(status_code=503, detail="serp_engine module not available.")

    suggestions = await _fetch_suggestions_intent(q.strip(), lang=lang)
    return {"keyword": q.strip(), "suggestions": suggestions}


@app.get("/intent")
def classify_crawled_keywords():
    """
    Classify all crawled page keywords by search intent.

    Reads from the current crawl_results (must crawl first).

    Response:
      {
        "page_count":    15,
        "distribution":  {"informational": 42, "commercial": 18, ...},
        "pages": [
          {
            "url":     "https://...",
            "keywords": [
              {"keyword": "seo guide", "intent": "informational",
               "intent_short": "Info", "intent_color": "#3b82f6"},
              ...
            ]
          }
        ]
      }
    """
    if not _INTENT_MODULE:
        raise HTTPException(status_code=503, detail="intent_classifier module not available.")

    from crawler import crawl_results as _cr
    if not _cr:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")

    total_dist = {"informational": 0, "commercial": 0,
                  "transactional": 0, "navigational": 0}
    pages_out  = []

    for page in _cr:
        if page.get("_is_error"):
            continue
        raw_kws = page.get("keywords") or []
        classified = []
        for k in raw_kws:
            kw = k if isinstance(k, str) else k.get("keyword", "")
            if not kw:
                continue
            intent = _classify_intent(kw)
            label  = _intent_label(intent)
            total_dist[intent] = total_dist.get(intent, 0) + 1
            classified.append({
                "keyword":      kw,
                "intent":       intent,
                "intent_short": label["short"],
                "intent_color": label["color"],
            })

        if classified:
            pages_out.append({
                "url":      page.get("url", ""),
                "keywords": classified,
            })

    return {
        "page_count":   len(pages_out),
        "distribution": total_dist,
        "pages":        pages_out,
    }


@app.get("/cannibalization")
def get_keyword_cannibalization():
    """
    Detect keyword cannibalization across all crawled pages.

    Returns pages on the same site that compete for the same keywords —
    these reduce each other's ranking potential.

    Reads from the current crawl_results (must crawl first).

    Response:
      {
        "total_conflicts": 3,
        "conflicts": [
          {
            "keyword":         "seo audit",
            "competing_pages": ["https://site.com/page-a", "https://site.com/page-b"],
            "page_count":      2,
            "risk_level":      "Medium",
            "recommendation":  "..."
          },
          ...
        ]
      }
    """
    if not _CANNIBAL_MODULE:
        raise HTTPException(status_code=503, detail="cannibalization module not available.")

    from crawler import crawl_results as _cr
    if not _cr:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")

    conflicts = _detect_cannibalization(list(_cr))
    return {
        "total_conflicts": len(conflicts),
        "conflicts":       conflicts,
    }


@app.get("/snippet-potential")
def get_snippet_potential(url: str = ""):
    """
    Score a crawled page's featured snippet potential.

    Query param:
      url — URL of a page already in crawl_results (required)

    Response:
      {
        "url":      "https://...",
        "score":    72,
        "potential": "High",
        "signals":  ["numbered_list", "question_headings"],
        "advice":   ["Add a 40-60 word definition paragraph after H1", ...]
      }
    """
    if not url.strip():
        raise HTTPException(status_code=400, detail="Query parameter 'url' is required.")
    if not _SERP_MODULE:
        raise HTTPException(status_code=503, detail="serp_engine module not available.")

    from crawler import crawl_results as _cr
    target_page = next(
        (p for p in _cr if p.get("url", "").rstrip("/") == url.strip().rstrip("/")),
        None,
    )
    if target_page is None:
        raise HTTPException(
            status_code=404,
            detail=f"URL not found in crawl results: {url}. Run /crawl first.",
        )

    result = _snippet_score(target_page)
    return {"url": url.strip(), **result}


# ════════════════════════════════════════════════════════════════════════════
# ── LINK GRAPH / PAGERANK / ORPHANS ─────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

@app.get("/link-graph")
def get_link_graph():
    """
    Full internal link graph analysis:
    PageRank scores, orphan pages, depth distribution, silo breakdown.
    Requires a completed crawl.
    """
    if not _LINK_GRAPH_MODULE:
        raise HTTPException(status_code=503, detail="link_graph module not available")
    pages = [p for p in crawl_results if not p.get("_is_error")]
    if not pages:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")
    return _analyse_link_graph(pages)


@app.get("/link-graph/orphans")
def get_orphan_pages():
    """Return only orphan pages (0 incoming internal links)."""
    if not _LINK_GRAPH_MODULE:
        raise HTTPException(status_code=503, detail="link_graph module not available")
    pages = [p for p in crawl_results if not p.get("_is_error")]
    if not pages:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")
    graph = _build_link_graph(pages)
    orphans = _detect_orphans(graph)
    return {"orphan_count": len(orphans), "orphans": orphans}


# ════════════════════════════════════════════════════════════════════════════
# ── CONTENT DEDUPLICATION ────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

@app.get("/content-dedup")
def get_content_dedup():
    """
    Near-duplicate content detection using SimHash.
    Returns page pairs that are ≥95% similar in body text.
    """
    if not _DEDUP_MODULE:
        raise HTTPException(status_code=503, detail="content_dedup module not available")
    pages = [p for p in crawl_results if not p.get("_is_error")]
    if not pages:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")
    # Pass body_text field (used by SimHash)
    dedup_pages = [{"url": p["url"], "body_text": p.get("body_text", "")} for p in pages]
    return _duplicate_summary(dedup_pages)


# ════════════════════════════════════════════════════════════════════════════
# ── SITE AUDITOR: robots / HSTS / mixed content / redirects ─────────────────
# ════════════════════════════════════════════════════════════════════════════

@app.get("/site-audit/full")
async def get_full_site_audit():
    """
    Full site-level audit:
    - robots.txt rules
    - HSTS header strength
    - Mixed content (HTTP resources on HTTPS pages)
    - Redirect chains for crawled URLs
    """
    if not _SITE_AUDITOR_MODULE:
        raise HTTPException(status_code=503, detail="site_auditor module not available")
    pages = [p for p in crawl_results if not p.get("_is_error")]
    if not pages:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")
    site_url = pages[0]["url"] if pages else ""
    return await _run_site_audit(site_url, pages)


@app.get("/site-audit/robots")
async def get_robots_audit(site_url: str = ""):
    """Fetch and parse /robots.txt for the crawled (or provided) site."""
    if not _SITE_AUDITOR_MODULE:
        raise HTTPException(status_code=503, detail="site_auditor module not available")
    if not site_url:
        pages = [p for p in crawl_results if not p.get("_is_error")]
        site_url = pages[0]["url"] if pages else ""
    if not site_url:
        raise HTTPException(status_code=400, detail="Provide site_url param or run /crawl first.")
    return await _fetch_robots_txt(site_url)


@app.get("/site-audit/mixed-content")
def get_mixed_content_audit():
    """Scan all crawled pages for mixed content (HTTP resources on HTTPS pages)."""
    if not _SITE_AUDITOR_MODULE:
        raise HTTPException(status_code=503, detail="site_auditor module not available")
    pages = [p for p in crawl_results if not p.get("_is_error")]
    if not pages:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")
    return _scan_mixed_content_all(pages)


@app.get("/site-audit/hsts")
def get_hsts_audit():
    """Check HSTS header on the first crawled page's response headers."""
    if not _SITE_AUDITOR_MODULE:
        raise HTTPException(status_code=503, detail="site_auditor module not available")
    pages = [p for p in crawl_results if not p.get("_is_error")]
    if not pages:
        raise HTTPException(status_code=400, detail="No crawl results. Run /crawl first.")
    # Use stored response_headers if available
    headers = pages[0].get("response_headers", {}) or {}
    return _check_hsts(headers)


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 7d: SERP POSITION + KEYWORD DIFFICULTY ENDPOINTS ────────────────
# ════════════════════════════════════════════════════════════════════════════

class SerpPositionRequest(BaseModel):
    keyword: str = Field(..., min_length=2, max_length=200)
    domain:  str = Field(..., min_length=4, max_length=2048)
    lang:    str = Field(default="en", min_length=2, max_length=5)
    num:     int = Field(default=30, ge=10, le=100)


class BulkSerpRequest(BaseModel):
    keywords: list[str] = Field(..., min_length=1, max_length=20)
    domain:   str       = Field(..., min_length=4, max_length=2048)


class DifficultyRequest(BaseModel):
    keywords: list[str] = Field(..., min_length=1, max_length=20)
    lang:     str       = Field(default="en", min_length=2, max_length=5)


@app.post("/serp/position")
async def check_serp_position(request: SerpPositionRequest):
    """
    Check where `domain` ranks for `keyword` on Google (top 30 results).

    Returns:
        {keyword, domain, position (1-30 or null), in_top_10, in_top_30}

    Note: scrapes Google HTML — use sparingly (1 req/call, ~2s delay built in).
    Returns 503 if serp_scraper module is unavailable.
    """
    if not _SERP_SCRAPER_MODULE:
        raise HTTPException(status_code=503, detail="serp_scraper module not available.")

    try:
        position = await _get_serp_position(
            request.keyword, request.domain,
            lang=request.lang, num=request.num,
        )
    except Exception as exc:
        logger.error("SERP position check failed for %r: %s", request.keyword, exc)
        raise HTTPException(status_code=500, detail="SERP check failed — see server logs.")

    return {
        "keyword":   request.keyword,
        "domain":    request.domain,
        "position":  position,
        "in_top_10": position is not None and position <= 10,
        "in_top_30": position is not None and position <= 30,
    }


@app.post("/serp/bulk-position")
async def bulk_serp_position(request: BulkSerpRequest):
    """
    Check SERP positions for multiple keywords for one domain (max 20).
    Requests are rate-limited (max 3 parallel) to avoid Google rate-limit.
    Returns [{keyword, position, in_top_10, in_top_30}]
    """
    if not _SERP_SCRAPER_MODULE:
        raise HTTPException(status_code=503, detail="serp_scraper module not available.")
    if len(request.keywords) > 20:
        raise HTTPException(status_code=400, detail="Maximum 20 keywords per request.")

    try:
        results = await _bulk_serp_check(request.keywords, request.domain)
    except Exception as exc:
        logger.error("Bulk SERP check failed: %s", exc)
        raise HTTPException(status_code=500, detail="Bulk SERP check failed — see server logs.")

    # Store positions on matching crawl pages so keyword scorer can emit
    # expected_ctr / ctr_tier on the next /pages or /ctr-opportunity call.
    pos_map: dict[str, int] = {
        r["keyword"]: r["position"]
        for r in results
        if r.get("position") is not None
    }
    if pos_map:
        norm_domain = request.domain.lower().rstrip("/").removeprefix("https://").removeprefix("http://").removeprefix("www.")
        for page in crawl_results:
            page_host = (page.get("url") or "").lower()
            if norm_domain in page_host:
                existing = page.get("serp_positions") or {}
                existing.update(pos_map)
                page["serp_positions"] = existing
                # Re-score keywords with new CTR data
                if page.get("keywords_scored") and not page.get("_is_error"):
                    try:
                        from keyword_scorer import score_keywords as _score_kw
                        suggest_set = set(page.get("_suggest_cache") or [])
                        page["keywords_scored"] = _score_kw(
                            page,
                            suggest_hits=suggest_set or None,
                            serp_positions=existing,
                        )
                    except Exception as _exc:
                        logger.warning("CTR re-score failed for %s: %s", page.get("url"), _exc)

    return {"domain": request.domain, "total": len(results), "results": results}


@app.post("/serp/difficulty")
async def keyword_difficulty(request: DifficultyRequest):
    """
    Estimate keyword difficulty for up to 20 keywords.

    Scrapes top-10 Google results per keyword, then fetches OPR scores
    for those domains (requires OPR_API_KEY env var for precise scores;
    falls back to result-count heuristic without it).

    Returns [{keyword, difficulty_score (0-100), difficulty_label, top_domains, avg_opr}]
    """
    if not _SERP_SCRAPER_MODULE:
        raise HTTPException(status_code=503, detail="serp_scraper module not available.")
    if len(request.keywords) > 20:
        raise HTTPException(status_code=400, detail="Maximum 20 keywords per request.")

    try:
        results = await _bulk_difficulty(request.keywords, concurrency=2)
    except Exception as exc:
        logger.error("Keyword difficulty check failed: %s", exc)
        raise HTTPException(status_code=500, detail="Difficulty check failed — see server logs.")

    return {"total": len(results), "results": results}


@app.get("/serp/visibility")
def serp_visibility():
    """
    Compute Visibility Score — % of crawled keywords ranked in top 10.

    Uses SERP positions stored by /serp/bulk-position calls.
    Returns per-keyword position data + aggregate visibility metrics.

    Response:
      {
        "total_keywords":   int,
        "in_top_3":         int,
        "in_top_10":        int,
        "in_top_30":        int,
        "not_ranked":       int,
        "visibility_score": float,  # weighted score 0-100
        "keywords":         [{"keyword", "position", "page_url", "expected_ctr"}]
      }
    """
    from serp_engine import expected_ctr as _ectr
    all_positions: list[dict] = []
    for page in crawl_results:
        pos_map  = page.get("serp_positions") or {}
        page_url = page.get("url", "")
        for kw, pos in pos_map.items():
            if isinstance(pos, int) and pos > 0:
                all_positions.append({
                    "keyword":      kw,
                    "position":     pos,
                    "page_url":     page_url,
                    "expected_ctr": _ectr(pos),
                })

    if not all_positions:
        return {
            "total_keywords":   0,
            "in_top_3":         0,
            "in_top_10":        0,
            "in_top_30":        0,
            "not_ranked":       0,
            "visibility_score": 0.0,
            "keywords":         [],
            "note": "Run POST /serp/bulk-position first to populate ranking data.",
        }

    total      = len(all_positions)
    top3       = sum(1 for k in all_positions if k["position"] <= 3)
    top10      = sum(1 for k in all_positions if k["position"] <= 10)
    top30      = sum(1 for k in all_positions if k["position"] <= 30)
    not_ranked = total - top30

    # Weighted visibility: top-3 worth 28.5%, top-10 avg 7.2%, top-30 avg 1.5%
    vis_score = round(
        (top3 * 28.5 + (top10 - top3) * 7.2 + (top30 - top10) * 1.5)
        / (total * 28.5) * 100,
        1,
    ) if total > 0 else 0.0

    all_positions.sort(key=lambda x: x["position"])
    return {
        "total_keywords":   total,
        "in_top_3":         top3,
        "in_top_10":        top10,
        "in_top_30":        top30,
        "not_ranked":       not_ranked,
        "visibility_score": min(vis_score, 100.0),
        "keywords":         all_positions,
    }


# ── CTR Opportunity endpoint ──────────────────────────────────────────────────

class CtrOpportunityRequest(BaseModel):
    """
    Calculate CTR uplift opportunities for a set of keyword→position pairs.

    ``positions`` maps each keyword to its current SERP rank (1-based).
    ``target_position`` is the rank you want to move each keyword to
    (default: 3, i.e. "what if I reach position 3?").
    ``intent`` is applied uniformly; pass "commercial" or "transactional"
    if the keyword set is clearly buy-intent.
    """
    positions:       dict[str, int] = Field(..., description="keyword → current position")
    target_position: int            = Field(default=3, ge=1, le=20)
    intent:          str            = Field(default="informational")


@app.post("/ctr-opportunity")
def ctr_opportunity(request: CtrOpportunityRequest):
    """
    Return CTR uplift analysis for each keyword given its current SERP position.

    Uses the Sistrix 2024 CTR benchmark curve built into serp_engine.py.
    No network calls — purely deterministic.

    Response:
      {
        "target_position": 3,
        "intent":          "informational",
        "results": [
          {
            "keyword":          "seo audit",
            "current_position": 8,
            "current_ctr":      0.032,
            "target_ctr":       0.11,
            "uplift_abs":       0.078,   # absolute CTR gain
            "uplift_rel":       3.4,     # relative multiplier
            "tier_change":      "bronze → gold",
          },
          ...
        ]
      }
    """
    from serp_engine import ctr_opportunity_score as _ctr_opportunity_score

    target  = request.target_position
    intent  = request.intent
    results = []

    for keyword, current_pos in request.positions.items():
        if not isinstance(current_pos, int) or current_pos < 1:
            continue
        opp = _ctr_opportunity_score(current_pos, target, intent)
        results.append({
            "keyword": keyword,
            **opp,
        })

    # Sort by absolute CTR uplift descending (biggest wins first)
    results.sort(key=lambda x: x.get("uplift_abs", 0), reverse=True)

    return {
        "target_position": target,
        "intent":          intent,
        "total":           len(results),
        "results":         results,
    }


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 8: SCHEDULED MONITORING ENDPOINTS (Phase 3) ─────────────────────
# ════════════════════════════════════════════════════════════════════════════

try:
    from monitor import (
        schedule_job   as _schedule_job,
        cancel_job     as _cancel_job,
        delete_job     as _delete_job,
        list_jobs      as _list_jobs,
        get_job        as _get_job,
        get_job_history as _get_job_history,
        get_domain_latest as _get_domain_latest,
        start_monitor_service as _start_monitor_service,
    )
    _MONITOR_MODULE = True
except ImportError:
    _MONITOR_MODULE = False


class MonitorScheduleRequest(BaseModel):
    domain:         str        = Field(..., min_length=4, max_length=2048)
    keywords:       list[str]  = Field(..., min_items=1, max_items=50)
    interval_hours: float      = Field(default=24.0, ge=0.5, le=168.0)


@app.post("/monitor/schedule")
async def monitor_schedule(request: MonitorScheduleRequest):
    """
    Schedule periodic SERP position tracking for a domain + keywords.

    The monitor fires every interval_hours, checks each keyword's position
    on Google, and saves results to the SQLite database.

    Returns the created MonitorJob with job_id for future management.
    """
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    if not _SERP_SCRAPER_MODULE:
        raise HTTPException(status_code=503, detail="serp_scraper module not available (needed for SERP checks).")

    job = _schedule_job(
        domain         = request.domain,
        keywords       = request.keywords,
        interval_hours = request.interval_hours,
    )
    return {"status": "scheduled", "job": job.__dict__ if hasattr(job, "__dict__") else job}


@app.get("/monitor/jobs")
def monitor_list_jobs():
    """List all active and inactive monitoring jobs."""
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    return {"total": len(_list_jobs()), "jobs": _list_jobs()}


@app.get("/monitor/job/{job_id}")
def monitor_get_job(job_id: str):
    """Get a single monitoring job by ID."""
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    job = _get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")
    return job


@app.delete("/monitor/job/{job_id}")
def monitor_delete_job(job_id: str):
    """Delete a monitoring job by ID."""
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    deleted = _delete_job(job_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Job not found.")
    return {"status": "deleted", "job_id": job_id}


@app.patch("/monitor/job/{job_id}/cancel")
def monitor_cancel_job(job_id: str):
    """Pause (deactivate) a monitoring job without deleting it."""
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    ok = _cancel_job(job_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Job not found.")
    return {"status": "cancelled", "job_id": job_id}


@app.get("/monitor/history")
def monitor_history(domain: str, keyword: str, limit: int = Query(default=30, ge=1, le=200)):
    """
    Retrieve position history for domain + keyword from the monitoring database.
    Returns newest-first list of {keyword, position, in_top_10, in_top_30, checked_at}.
    """
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    results = _get_job_history(domain, keyword, limit=limit)
    return {"domain": domain, "keyword": keyword, "total": len(results), "history": results}


@app.get("/monitor/latest")
def monitor_latest(domain: str):
    """Return the latest tracked position for every keyword on domain."""
    if not _MONITOR_MODULE:
        raise HTTPException(status_code=503, detail="monitor module not available.")
    results = _get_domain_latest(domain)
    return {"domain": domain, "total": len(results), "results": results}


# ── Start monitor service on FastAPI startup ──────────────────────────────────

@app.on_event("startup")
async def _startup_monitor():
    """Start the background SERP monitor loop when FastAPI boots."""
    if _MONITOR_MODULE:
        try:
            _start_monitor_service()
            logger.info("Monitor service started at FastAPI startup")
        except Exception as exc:
            logger.warning("Monitor service startup failed: %s", exc)


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 9: PDF EXPORT ENDPOINT (Phase 4) ────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

try:
    from pdf_export import generate_pdf_bytes as _generate_pdf_bytes
    _PDF_MODULE = True
except ImportError:
    _PDF_MODULE = False


@app.get("/export-pdf")
def export_pdf(url: str = Query(default="", description="Crawled site URL for report header")):
    """
    Generate and download a PDF audit report of the last crawl.

    Requires reportlab to be installed (pip install reportlab>=4.0.0).
    Returns a PDF file with:
      - Cover header (site URL, crawl date, stats)
      - Summary metrics (pages, issues, high priority, clean)
      - Issue breakdown table
      - Per-page audit table (top 100 pages, sorted by priority)
      - Top keywords
    """
    if not _PDF_MODULE:
        raise HTTPException(
            status_code=503,
            detail="PDF export unavailable — add 'reportlab>=4.0.0' to requirements.txt.",
        )
    if not crawl_results:
        raise HTTPException(status_code=400, detail="No crawl data. Run a crawl first.")

    site_url = url.strip() or (
        crawl_results[0].get("url", "") if crawl_results else ""
    )
    try:
        pdf_bytes = _generate_pdf_bytes(
            list(crawl_results),
            dict(crawl_status),
            site_url=site_url,
        )
    except Exception as exc:
        logger.error("PDF generation failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {exc}")

    from fastapi.responses import Response as _Resp
    from urllib.parse import quote as _quote
    safe_host = (site_url or "crawliq").replace("https://", "").replace("http://", "").rstrip("/").replace("/", "_")[:40]
    filename  = f"crawliq_audit_{safe_host}.pdf"
    return _Resp(
        content      = pdf_bytes,
        media_type   = "application/pdf",
        headers      = {"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ════════════════════════════════════════════════════════════════════════════
# ── SECTION 10: SaaS ENDPOINTS (Phase 5) ─────────────────────────────────────
#   Auth, Projects, Issue Status, Keyword Gap, API Keys, Sitemap Crawl,
#   Score History, User Settings (white-label logo, alert prefs)
# ════════════════════════════════════════════════════════════════════════════

# ── Auth module (graceful fallback) ──────────────────────────────────────────
try:
    import auth as _auth_mod
    from auth import (
        register                  as _auth_register,
        login                     as _auth_login,
        get_user_by_token         as _auth_by_token,
        get_user_by_api_key       as _auth_by_key,
        get_user_by_id            as _auth_by_id,
        update_user               as _auth_update,
        rotate_api_key            as _auth_rotate_key,
        check_crawl_quota         as _auth_check_quota,
        record_pages_crawled      as _auth_record_pages,
        create_password_reset_token as _auth_create_reset_token,
        reset_password            as _auth_reset_password,
        create_email_verify_token as _auth_create_verify_token,
        verify_email_token        as _auth_verify_email,
        TIER_LIMITS               as _TIER_LIMITS,
    )
    _AUTH_MODULE = True
except ImportError:
    _AUTH_MODULE = False

# ── Billing module (graceful fallback — requires stripe package) ──────────────
try:
    import billing as _billing_mod
    from billing import (
        is_configured           as _billing_configured,
        create_checkout_session as _billing_checkout,
        create_portal_session   as _billing_portal,
        handle_webhook          as _billing_webhook,
        get_subscription_status as _billing_status,
    )
    _BILLING_MODULE = True
except ImportError:
    _BILLING_MODULE = False

# ── DB project / snapshot / issue helpers ────────────────────────────────────
try:
    from competitor_db import (
        create_project      as _db_create_project,
        list_projects       as _db_list_projects,
        get_project         as _db_get_project,
        update_project      as _db_update_project,
        delete_project      as _db_delete_project,
        save_crawl_snapshot as _db_save_snapshot,
        get_crawl_history   as _db_get_history,
        upsert_issue_status as _db_upsert_issue,
        get_issue_statuses  as _db_get_issues,
    )
    _PROJECTS_MODULE = True
except ImportError:
    _PROJECTS_MODULE = False


# ── Auth helpers ──────────────────────────────────────────────────────────────

def _get_current_user(request) -> dict | None:
    """
    Extract user from Authorization: Bearer <token> header or X-API-Key header.
    Returns user dict or None (unauthenticated).
    """
    if not _AUTH_MODULE:
        return None
    auth_header = request.headers.get("Authorization", "")
    api_key_header = request.headers.get("X-API-Key", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:].strip()
        return _auth_by_token(token)
    if api_key_header:
        return _auth_by_key(api_key_header)
    return None


# ── Pydantic models ───────────────────────────────────────────────────────────

import base64 as _base64

class RegisterRequest(BaseModel):
    email:    str = Field(..., min_length=5, max_length=254)
    password: str = Field(..., min_length=6, max_length=128)
    name:     str = Field(default="", max_length=100)

class LoginRequest(BaseModel):
    email:    str
    password: str

class ProjectCreateRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    url:  str = Field(..., min_length=4, max_length=2048)

class IssueStatusRequest(BaseModel):
    project_id: int | None = None
    url:        str
    issue_type: str
    status:     str = Field(..., pattern="^(open|in_progress|resolved)$")
    note:       str = ""

class UserSettingsRequest(BaseModel):
    name:                str | None = None
    alert_email:         str | None = None
    rank_drop_threshold: int | None = None

class KeywordGapRequest(BaseModel):
    your_keywords:       list[str] = Field(..., max_items=500)
    competitor_keywords: list[str] = Field(..., max_items=500)

class BillingCheckoutRequest(BaseModel):
    tier: str = Field(..., pattern="^(pro|agency)$")

class ForgotPasswordRequest(BaseModel):
    email: str = Field(..., min_length=5, max_length=254)

class ResetPasswordRequest(BaseModel):
    token:        str = Field(..., min_length=10)
    new_password: str = Field(..., min_length=6, max_length=128)


# ── Auth endpoints ────────────────────────────────────────────────────────────

@app.post("/auth/register")
@_rate_limit("5/minute")
def auth_register(req: RegisterRequest, request: _FastAPIRequest):
    """Register a new user. Returns access token."""
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available (install python-jose + passlib)")
    try:
        user  = _auth_register(req.email, req.password, req.name)
        token = _auth_login(req.email, req.password)
        # Send email verification token (best-effort)
        try:
            verify_token = _auth_create_verify_token(user["id"])
            from email_alerts import send_email_verify
            import asyncio as _asyncio
            _asyncio.create_task(send_email_verify(user["email"], verify_token))
        except Exception:
            pass
        return {"token": token, "user": user}
    except ValueError as exc:
        raise HTTPException(400, str(exc))


@app.post("/auth/login")
@_rate_limit("5/minute")
def auth_login(req: LoginRequest, request: _FastAPIRequest):
    """Login and return JWT access token."""
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available")
    try:
        token = _auth_login(req.email, req.password)
        user  = _auth_by_token(token)
        return {"token": token, "user": user}
    except ValueError as exc:
        raise HTTPException(401, str(exc))


@app.get("/auth/me")
def auth_me(request: _FastAPIRequest):
    """Return current user profile (requires Bearer token)."""
    user = _get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    # Enrich with tier limits
    tier   = user.get("tier", "free")
    limits = (_TIER_LIMITS if _AUTH_MODULE else {}).get(tier, {})
    return {**user, "limits": limits}


@app.post("/auth/api-key/rotate")
def auth_rotate_api_key(request: _FastAPIRequest):
    """Generate a new API key for the current user."""
    user = _get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    new_key = _auth_rotate_key(user["id"])
    return {"api_key": new_key}


@app.post("/auth/forgot-password")
@_rate_limit("3/minute")
def auth_forgot_password(req: ForgotPasswordRequest, request: _FastAPIRequest):
    """
    Initiate password reset. Sends a reset link to the registered email.
    Always returns 200 to avoid leaking whether the email is registered.
    """
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available")
    token = _auth_create_reset_token(req.email)
    if token:
        try:
            from email_alerts import send_password_reset
            import asyncio as _asyncio
            _asyncio.create_task(send_password_reset(req.email, token))
        except Exception:
            pass
    return {"status": "ok", "message": "If that email is registered, a reset link has been sent."}


@app.post("/auth/reset-password")
def auth_reset_password(req: ResetPasswordRequest):
    """Consume a password reset token and set a new password."""
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available")
    try:
        ok = _auth_reset_password(req.token, req.new_password)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    if not ok:
        raise HTTPException(400, "Invalid or expired reset token.")
    return {"status": "ok", "message": "Password updated. You can now log in."}


@app.get("/auth/verify-email/{token}")
def auth_verify_email(token: str):
    """Verify email address via token sent at registration."""
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available")
    ok = _auth_verify_email(token)
    if not ok:
        raise HTTPException(400, "Invalid or expired verification link.")
    return {"status": "ok", "message": "Email verified."}


# ── Billing endpoints ─────────────────────────────────────────────────────────

@app.post("/billing/checkout")
def billing_checkout(req: BillingCheckoutRequest, request: _FastAPIRequest):
    """Create a Stripe Checkout session to upgrade to pro or agency tier."""
    if not _BILLING_MODULE:
        raise HTTPException(503, "Billing module not available (install stripe>=7.0.0)")
    user = _get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    try:
        result = _billing_checkout(user, req.tier)
        return result
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    except RuntimeError as exc:
        raise HTTPException(503, str(exc))
    except Exception as exc:
        logger.error("Billing checkout error: %s", exc)
        raise HTTPException(500, "Checkout session creation failed")


@app.post("/billing/webhook")
async def billing_webhook(request: _FastAPIRequest):
    """
    Receive and process Stripe webhook events.
    Stripe signs each payload with a webhook secret — always verify.
    """
    if not _BILLING_MODULE:
        raise HTTPException(503, "Billing module not available")
    payload = await request.body()
    sig = request.headers.get("stripe-signature", "")
    try:
        result = _billing_webhook(payload, sig)
        return result
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    except Exception as exc:
        logger.error("Billing webhook error: %s", exc)
        raise HTTPException(500, "Webhook processing failed")


@app.get("/billing/portal")
def billing_portal(request: _FastAPIRequest):
    """Create a Stripe Customer Portal session for managing subscription."""
    if not _BILLING_MODULE:
        raise HTTPException(503, "Billing module not available")
    user = _get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    customer_id = user.get("stripe_customer_id")
    if not customer_id:
        raise HTTPException(400, "No active subscription found. Please upgrade first.")
    try:
        return _billing_portal(customer_id)
    except Exception as exc:
        logger.error("Billing portal error: %s", exc)
        raise HTTPException(500, "Portal session creation failed")


@app.get("/billing/status")
def billing_status(request: _FastAPIRequest):
    """Return current subscription status for the authenticated user."""
    user = _get_current_user(request)
    if not user:
        return {"tier": "free", "status": "unauthenticated"}
    customer_id = user.get("stripe_customer_id")
    if not _BILLING_MODULE or not customer_id:
        return {"tier": user.get("tier", "free"), "status": "local"}
    try:
        return _billing_status(customer_id)
    except Exception:
        return {"tier": user.get("tier", "free"), "status": "error"}


@app.patch("/user/settings")
def user_settings(req: UserSettingsRequest, request: _FastAPIRequest):
    """Update user profile: display name, alert email, rank-drop threshold."""
    user = _get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    kwargs = {k: v for k, v in req.dict().items() if v is not None}
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available")
    _auth_update(user["id"], **kwargs)
    return {"status": "updated"}


@app.post("/user/logo")
async def user_upload_logo(request: _FastAPIRequest, file: UploadFile = File(...)):
    """
    Upload a logo image (PNG/JPG, max 512 KB) stored as base64 for white-label PDF.
    """
    user = _get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    content = await file.read()
    if len(content) > 512 * 1024:
        raise HTTPException(400, "Logo file too large (max 512 KB)")
    if not content[:4] in (b"\x89PNG", b"\xff\xd8\xff"):
        pass   # allow other image types
    encoded = _base64.b64encode(content).decode()
    if _AUTH_MODULE:
        _auth_update(user["id"], logo_base64=encoded)
    return {"status": "uploaded", "size": len(content)}


# ── Project endpoints ─────────────────────────────────────────────────────────

@app.get("/projects")
def list_projects(request: _FastAPIRequest):
    """List all projects for the current user (or all if no auth)."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    user = _get_current_user(request)
    uid  = user["id"] if user else None
    return {"projects": _db_list_projects(uid)}


@app.post("/projects")
def create_project(req: ProjectCreateRequest, request: _FastAPIRequest):
    """Create a new project (named crawl session)."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    user = _get_current_user(request)
    uid  = user["id"] if user else None
    proj = _db_create_project(uid, req.name, req.url)
    return {"project": proj}


@app.get("/projects/{project_id}")
def get_project(project_id: int, request: _FastAPIRequest):
    """Get a single project by ID."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    proj = _db_get_project(project_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    # Ownership check: only owner or project members may view
    user = _get_current_user(request)
    if user and proj.get("user_id") and proj["user_id"] != user["id"]:
        # Check if user is a member
        try:
            from competitor_db import _connect as _dbc
            with _dbc() as conn:
                member = conn.execute(
                    "SELECT 1 FROM project_members WHERE project_id=? AND user_id=?",
                    (project_id, user["id"])
                ).fetchone()
            if not member:
                raise HTTPException(403, "Access denied")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(403, "Access denied")
    return proj


@app.delete("/projects/{project_id}")
def delete_project(project_id: int, request: _FastAPIRequest):
    """Delete a project and all its snapshots. Only the owner may delete."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    proj = _db_get_project(project_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    user = _get_current_user(request)
    if user and proj.get("user_id") and proj["user_id"] != user["id"]:
        raise HTTPException(403, "Only the project owner can delete this project")
    deleted = _db_delete_project(project_id)
    if not deleted:
        raise HTTPException(404, "Project not found")
    return {"status": "deleted", "project_id": project_id}


@app.post("/projects/{project_id}/snapshot")
def save_snapshot(project_id: int, request: _FastAPIRequest):
    """
    Save a crawl snapshot for the given project.
    Reads from the current in-memory crawl_results + crawl_status.
    """
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    if not crawl_results:
        raise HTTPException(400, "No crawl data to snapshot. Run a crawl first.")
    proj = _db_get_project(project_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    # Only owner or editors may save snapshots
    user = _get_current_user(request)
    if user and proj.get("user_id") and proj["user_id"] != user["id"]:
        try:
            from competitor_db import _connect as _dbc
            with _dbc() as conn:
                member = conn.execute(
                    "SELECT role FROM project_members WHERE project_id=? AND user_id=?",
                    (project_id, user["id"])
                ).fetchone()
            if not member or member["role"] not in ("editor",):
                raise HTTPException(403, "Editor or owner access required to save snapshots")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(403, "Access denied")

    pages       = [p for p in crawl_results if not p.get("_is_error")]
    issue_count = sum(1 for p in pages if p.get("issues"))
    scores      = [p.get("seo_score", 0) for p in pages if p.get("seo_score") is not None]
    health      = round(sum(scores) / len(scores), 1) if scores else 0

    # Lightweight results json: top 100 pages, no full HTML
    top_pages = sorted(pages, key=lambda p: p.get("seo_score", 0))[:100]
    light = [{
        "url":    p.get("url", ""),
        "title":  (p.get("title") or "")[:100],
        "score":  p.get("seo_score"),
        "issues": p.get("issues", []),
        "priority": p.get("priority", ""),
    } for p in top_pages]

    snap_id = _db_save_snapshot(
        project_id   = project_id,
        page_count   = len(pages),
        issue_count  = issue_count,
        health_score = health,
        results_json = json.dumps(light),
    )
    _db_update_project(
        project_id,
        page_count   = len(pages),
        issue_count  = issue_count,
        health_score = health,
        last_crawl_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    )
    return {"snapshot_id": snap_id, "health_score": health,
            "page_count": len(pages), "issue_count": issue_count}


@app.get("/projects/{project_id}/history")
def get_project_history(project_id: int):
    """Return score history for a project (for the health-score trend chart)."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    history = _db_get_history(project_id)
    return {"project_id": project_id, "history": history}


# ── Issue status endpoints ────────────────────────────────────────────────────

@app.patch("/issues/status")
def update_issue_status(req: IssueStatusRequest, request: _FastAPIRequest):
    """Mark an issue as open / in_progress / resolved, with optional note."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    # Auth required when project_id is provided
    if req.project_id is not None and _AUTH_MODULE:
        user = _get_current_user(request)
        if not user:
            raise HTTPException(401, "Authentication required to update issue status")
        proj = _db_get_project(req.project_id)
        if proj and proj.get("user_id") and proj["user_id"] != user["id"]:
            try:
                from competitor_db import _connect as _dbc
                with _dbc() as conn:
                    member = conn.execute(
                        "SELECT role FROM project_members WHERE project_id=? AND user_id=?",
                        (req.project_id, user["id"])
                    ).fetchone()
                if not member:
                    raise HTTPException(403, "Access denied")
            except HTTPException:
                raise
            except Exception:
                raise HTTPException(403, "Access denied")
    _db_upsert_issue(req.project_id, req.url, req.issue_type, req.status, req.note)
    return {"status": "updated"}


@app.get("/issues/status")
def get_issue_statuses_endpoint(
    project_id: int | None = Query(default=None),
    url:        str | None = Query(default=None),
):
    """Get all issue statuses for a project (optionally filtered by URL)."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    rows = _db_get_issues(project_id, url)
    return {"statuses": rows}


# ── Keyword gap endpoint ──────────────────────────────────────────────────────

@app.post("/keyword-gap")
def keyword_gap(req: KeywordGapRequest):
    """
    Compare two keyword sets and return gap analysis.

    Returns:
      - only_competitor: keywords competitor has that you don't
      - only_you: keywords you have that competitor doesn't
      - shared: keywords both rank for
    """
    yours = {k.strip().lower() for k in req.your_keywords if k.strip()}
    theirs = {k.strip().lower() for k in req.competitor_keywords if k.strip()}

    only_comp = sorted(theirs - yours)
    only_you  = sorted(yours - theirs)
    shared    = sorted(yours & theirs)

    return {
        "your_total":         len(yours),
        "competitor_total":   len(theirs),
        "only_competitor":    only_comp,
        "only_you":           only_you,
        "shared":             shared,
        "gap_count":          len(only_comp),
        "opportunity_count":  len(only_comp),
    }


# ── Sitemap-driven crawl endpoint ─────────────────────────────────────────────

@app.post("/sitemap-crawl")
async def sitemap_crawl(
    sitemap_url:  str   = Query(..., description="URL of the sitemap.xml"),
    max_pages:    int   = Query(default=100, ge=1, le=500),
    background_tasks: BackgroundTasks = None,
):
    """
    Parse a sitemap.xml and crawl only those URLs (not BFS from root).
    Returns immediately; results available via /results and /crawl-status.
    """
    if crawl_status.get("running"):
        raise HTTPException(400, "A crawl is already running. Wait for it to finish.")

    # SSRF guard — same blocklist as /crawl
    _sm_parsed = _urlparse(sitemap_url)
    _sm_host   = (_sm_parsed.hostname or "").lower()
    if _sm_host in _SSRF_BLOCKED or _sm_host.startswith(("192.168.", "10.", "172.16.", "169.254.")):
        raise HTTPException(400, "Private or reserved IP addresses are not allowed.")

    # Fetch and parse sitemap
    urls = []
    try:
        import aiohttp as _aio
        async with _aio.ClientSession() as _sess:
            async with _sess.get(sitemap_url, timeout=_aio.ClientTimeout(total=15), ssl=False) as resp:
                if resp.status != 200:
                    raise HTTPException(400, f"Sitemap fetch failed (HTTP {resp.status})")
                xml_text = await resp.text(errors="replace")
        import re as _re
        urls = _re.findall(r"<loc>\s*(https?://[^<\s]+)\s*</loc>", xml_text, _re.IGNORECASE)
        urls = list(dict.fromkeys(urls))[:max_pages]   # dedup + cap
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Sitemap parse error: {exc}")

    if not urls:
        raise HTTPException(400, "No <loc> URLs found in sitemap.")

    # Fire a synthetic crawl using the first URL as root, seed queue with all sitemap URLs
    root_url = urls[0]
    crawl_results.clear()
    crawl_status.update({
        "running": True, "done": False, "error": None,
        "pages_crawled": 0, "total_pages": len(urls),
        "current_url": "", "elapsed_s": 0,
        "source": "sitemap",
    })

    async def _sitemap_crawl_task():
        import time as _time
        start = _time.time()
        from crawler import SEOCrawler
        try:
            crawler = SEOCrawler(root_url, max_pages=len(urls))
            # Seed the crawler's queue with sitemap URLs
            crawler.queue = [(u, 1) for u in urls]
            crawler._queued = set(urls)
            results = await crawler.crawl()
            crawl_results.extend(results)
            crawl_status.update({
                "running": False, "done": True,
                "pages_crawled": len(results),
                "elapsed_s": round(_time.time() - start, 1),
            })
        except Exception as exc:
            crawl_status.update({"running": False, "done": False, "error": str(exc)})

    if background_tasks:
        background_tasks.add_task(_sitemap_crawl_task)
    else:
        asyncio.create_task(_sitemap_crawl_task())

    return {"status": "started", "urls_found": len(urls), "root": root_url}


# ── Usage / quota endpoint ────────────────────────────────────────────────────

@app.get("/user/usage")
def user_usage(request: _FastAPIRequest):
    """Return current user's crawl credit usage and tier limits."""
    if not _AUTH_MODULE:
        return {"tier": "unlimited", "pages_used": 0, "pages_limit": -1}
    user = _get_current_user(request)
    if not user:
        return {"tier": "free", "pages_used": 0, "pages_limit": 200}
    tier   = user.get("tier", "free")
    limits = _TIER_LIMITS.get(tier, _TIER_LIMITS["free"])
    return {
        "tier":         tier,
        "pages_used":   user.get("pages_used", 0),
        "pages_limit":  limits["pages_per_month"],
        "projects_limit": limits["projects"],
        "monitor_limit":  limits["monitor_jobs"],
        "pages_reset_at": user.get("pages_reset_at"),
    }


# ── Update monitor/schedule to accept alert settings ─────────────────────────

class MonitorScheduleRequestV2(BaseModel):
    domain:          str       = Field(..., min_length=4, max_length=2048)
    keywords:        list[str] = Field(..., min_items=1, max_items=50)
    interval_hours:  float     = Field(default=24.0, ge=0.5, le=168.0)
    alert_email:     str | None = None
    drop_threshold:  int       = Field(default=5, ge=1, le=100)


@app.post("/monitor/schedule/v2")
async def monitor_schedule_v2(request: MonitorScheduleRequestV2):
    """Schedule monitoring with email alert support."""
    if not _MONITOR_MODULE:
        raise HTTPException(503, "monitor module not available")
    if not _SERP_SCRAPER_MODULE:
        raise HTTPException(503, "serp_scraper module not available")
    from monitor import schedule_job as _sched
    job = _sched(
        domain         = request.domain,
        keywords       = request.keywords,
        interval_hours = request.interval_hours,
        alert_email    = request.alert_email,
        drop_threshold = request.drop_threshold,
    )
    import dataclasses
    return {"status": "scheduled", "job": dataclasses.asdict(job)}


# ── White-label PDF endpoint ──────────────────────────────────────────────────

@app.get("/export-pdf/branded")
def export_pdf_branded(
    request:    _FastAPIRequest,
    url:        str = Query(default=""),
    brand_name: str = Query(default=""),
):
    """
    Generate PDF with the user's brand name and logo (white-label).
    Falls back to standard /export-pdf if not authenticated.
    """
    if not _PDF_MODULE:
        raise HTTPException(503, "PDF module not available")
    if not crawl_results:
        raise HTTPException(400, "No crawl data. Run a crawl first.")

    user = _get_current_user(request)
    effective_brand = brand_name.strip() or (
        user.get("name", "") if user else ""
    ) or "CrawlIQ"

    site_url = url.strip() or (
        crawl_results[0].get("url", "") if crawl_results else ""
    )
    try:
        pdf_bytes = _generate_pdf_bytes(
            list(crawl_results),
            dict(crawl_status),
            site_url   = site_url,
            brand_name = effective_brand,
        )
    except Exception as exc:
        raise HTTPException(500, f"PDF generation failed: {exc}")

    safe_brand = effective_brand.replace(" ", "_")[:20]
    safe_host  = (site_url or "site").replace("https://", "").replace("http://", "").rstrip("/")[:30]
    filename   = f"{safe_brand}_audit_{safe_host}.pdf"
    return Response(
        content    = pdf_bytes,
        media_type = "application/pdf",
        headers    = {"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Google Search Console OAuth + Data ───────────────────────────────────────
try:
    from google_auth_oauthlib.flow import Flow as _GscFlow
    from google.oauth2.credentials import Credentials as _GscCreds
    import google.auth.transport.requests as _GscTransport
    from googleapiclient.discovery import build as _gsc_build
    _GSC_AVAILABLE = True
except ImportError:
    _GSC_AVAILABLE = False

_GSC_TOKEN_FILE = os.path.join(os.path.dirname(__file__), "gsc_token.json")
_GSC_SCOPES     = ["https://www.googleapis.com/auth/webmasters.readonly"]


def _gsc_client_config() -> dict:
    redirect = os.getenv(
        "GSC_REDIRECT_URI",
        "https://bhavani7-seo-project.hf.space/gsc/callback",
    )
    # GSC_CLIENT_ID may be stored as just the prefix (without .apps.googleusercontent.com)
    # to avoid HF Spaces secret validation rejecting values with dots.
    _raw_id = os.getenv("GSC_CLIENT_ID", "")
    client_id = (
        _raw_id if ".apps.googleusercontent.com" in _raw_id
        else _raw_id + ".apps.googleusercontent.com"
    )
    return {
        "web": {
            "client_id":     client_id,
            "client_secret": os.getenv("GSC_CLIENT_SECRET", ""),
            "redirect_uris": [redirect],
            "auth_uri":      "https://accounts.google.com/o/oauth2/auth",
            "token_uri":     "https://oauth2.googleapis.com/token",
        }
    }


def _gsc_load_creds():
    if not _GSC_AVAILABLE or not os.path.exists(_GSC_TOKEN_FILE):
        return None
    try:
        with open(_GSC_TOKEN_FILE) as f:
            data = json.load(f)
        _raw_id = os.getenv("GSC_CLIENT_ID", "")
        _full_id = (
            _raw_id if ".apps.googleusercontent.com" in _raw_id
            else _raw_id + ".apps.googleusercontent.com"
        )
        creds = _GscCreds(
            token=data.get("token"),
            refresh_token=data.get("refresh_token"),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=_full_id,
            client_secret=os.getenv("GSC_CLIENT_SECRET"),
            scopes=_GSC_SCOPES,
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(_GscTransport.Request())
            _gsc_save_creds(creds)
        return creds
    except Exception:
        return None


def _gsc_save_creds(creds) -> None:
    with open(_GSC_TOKEN_FILE, "w") as f:
        json.dump({"token": creds.token, "refresh_token": creds.refresh_token}, f)


@app.get("/gsc/auth-url")
def gsc_auth_url():
    """Return Google OAuth2 URL for Search Console. Requires GSC_CLIENT_ID + GSC_CLIENT_SECRET."""
    client_id     = os.getenv("GSC_CLIENT_ID", "")
    client_secret = os.getenv("GSC_CLIENT_SECRET", "")
    if not client_id or not client_secret:
        return {
            "available": False,
            "message": "Set GSC_CLIENT_ID and GSC_CLIENT_SECRET env vars to enable Search Console.",
        }
    if not _GSC_AVAILABLE:
        return {
            "available": False,
            "message": "Run: pip install google-auth google-auth-oauthlib google-api-python-client",
        }
    redirect = os.getenv("GSC_REDIRECT_URI", "http://localhost:7860/gsc/callback")
    flow = _GscFlow.from_client_config(_gsc_client_config(), scopes=_GSC_SCOPES)
    flow.redirect_uri = redirect
    auth_url, _ = flow.authorization_url(access_type="offline", prompt="consent")
    return {"available": True, "auth_url": auth_url}


@app.get("/gsc/callback")
async def gsc_callback(
    code:  str = Query(default=""),
    error: str = Query(default=""),
):
    """Exchange OAuth code for access + refresh tokens and persist them."""
    if error:
        raise HTTPException(400, f"OAuth error: {error}")
    if not code:
        raise HTTPException(400, "No OAuth code received")
    if not _GSC_AVAILABLE:
        raise HTTPException(503, "GSC libraries not installed — run pip install google-auth-oauthlib google-api-python-client")
    redirect = os.getenv("GSC_REDIRECT_URI", "http://localhost:7860/gsc/callback")
    try:
        flow = _GscFlow.from_client_config(_gsc_client_config(), scopes=_GSC_SCOPES)
        flow.redirect_uri = redirect
        flow.fetch_token(code=code)
        _gsc_save_creds(flow.credentials)
    except Exception as exc:
        raise HTTPException(500, f"Token exchange failed: {exc}")
    return HTMLResponse("""
    <html><body style="background:#0d0d0d;color:#fff;font-family:monospace;padding:40px;text-align:center">
    <h2 style="color:#22c55e">&#10003; Connected to Google Search Console</h2>
    <p style="color:#888">You can close this window. CrawlIQ will now pull real impressions and CTR data.</p>
    <script>
      if(window.opener){window.opener.postMessage({gsc:'connected'},'*');}
      setTimeout(()=>window.close(), 2000);
    </script>
    </body></html>
    """)


@app.get("/gsc/status")
def gsc_status():
    """Return whether GSC is connected and tokens are valid."""
    if not _GSC_AVAILABLE:
        return {"connected": False, "reason": "libraries_missing"}
    creds = _gsc_load_creds()
    if not creds or not creds.valid:
        return {"connected": False, "reason": "not_authenticated"}
    return {"connected": True}


@app.delete("/gsc/disconnect")
def gsc_disconnect():
    """Remove stored GSC tokens."""
    if os.path.exists(_GSC_TOKEN_FILE):
        os.remove(_GSC_TOKEN_FILE)
    return {"disconnected": True}


@app.get("/gsc/sites")
def gsc_sites():
    """List all GSC-verified properties for the connected account."""
    if not _GSC_AVAILABLE:
        raise HTTPException(503, "GSC libraries not installed")
    creds = _gsc_load_creds()
    if not creds:
        raise HTTPException(401, "Not authenticated with GSC")
    try:
        svc   = _gsc_build("webmasters", "v3", credentials=creds, cache_discovery=False)
        resp  = svc.sites().list().execute()
        sites = [s["siteUrl"] for s in resp.get("siteEntry", [])]
        return {"sites": sites}
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.get("/gsc/data")
async def gsc_data(
    site_url: str = Query(..., description="GSC property URL e.g. https://example.com/"),
    days:     int = Query(default=28, ge=7, le=90),
):
    """
    Fetch clicks, impressions, CTR, and avg-position from Search Console.
    Returns summary totals + top-25 keywords for the given date range.
    GSC data has a ~3-day lag so end date is today-3.
    """
    if not _GSC_AVAILABLE:
        raise HTTPException(503, "GSC libraries not installed")
    creds = _gsc_load_creds()
    if not creds:
        raise HTTPException(401, "Not authenticated with GSC — connect Search Console first")
    import datetime
    end   = datetime.date.today() - datetime.timedelta(days=3)
    start = end - datetime.timedelta(days=days)
    try:
        svc  = _gsc_build("webmasters", "v3", credentials=creds, cache_discovery=False)
        body = {
            "startDate":  start.isoformat(),
            "endDate":    end.isoformat(),
            "dimensions": ["query"],
            "rowLimit":   25,
        }
        resp = svc.searchAnalytics().query(siteUrl=site_url, body=body).execute()
        rows = resp.get("rows", [])
        total_clicks = sum(r["clicks"]      for r in rows)
        total_impr   = sum(r["impressions"] for r in rows)
        avg_ctr      = (total_clicks / total_impr * 100) if total_impr else 0.0
        avg_pos      = (sum(r["position"] for r in rows) / len(rows)) if rows else 0.0
        keywords = [
            {
                "keyword":     r["keys"][0],
                "clicks":      r["clicks"],
                "impressions": r["impressions"],
                "ctr":         round(r["ctr"] * 100, 2),
                "position":    round(r["position"], 1),
            }
            for r in rows
        ]
        return {
            "site_url":   site_url,
            "date_range": {"start": start.isoformat(), "end": end.isoformat()},
            "summary": {
                "clicks":       total_clicks,
                "impressions":  total_impr,
                "ctr":          round(avg_ctr, 2),
                "avg_position": round(avg_pos, 1),
            },
            "top_keywords": keywords,
        }
    except Exception as exc:
        raise HTTPException(500, str(exc))


# ── Crawl diff endpoint ───────────────────────────────────────────────────────

@app.get("/projects/{project_id}/diff")
def project_diff(project_id: int, request: _FastAPIRequest):
    """
    Compare the two most recent crawl snapshots for a project.
    Returns lists of new_issues (appeared), fixed_issues (resolved),
    and score_delta.
    """
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    proj = _db_get_project(project_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    user = _get_current_user(request)
    if user and proj.get("user_id") and proj["user_id"] != user["id"]:
        try:
            from competitor_db import _connect as _dbc
            with _dbc() as conn:
                member = conn.execute(
                    "SELECT 1 FROM project_members WHERE project_id=? AND user_id=?",
                    (project_id, user["id"])
                ).fetchone()
            if not member:
                raise HTTPException(403, "Access denied")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(403, "Access denied")
    history = _db_get_history(project_id, limit=2)
    if len(history) < 2:
        return {"has_diff": False, "message": "Need at least 2 snapshots to compare. Run another crawl and save."}
    newer, older = history[0], history[1]

    # Parse stored results_json from both snapshots
    try:
        from competitor_db import _connect as _db_conn
        with _db_conn() as conn:
            rows = conn.execute(
                "SELECT results_json, crawled_at FROM crawl_snapshots WHERE project_id=? ORDER BY crawled_at DESC LIMIT 2",
                (project_id,)
            ).fetchall()
        if len(rows) < 2:
            return {"has_diff": False, "message": "Snapshot data not available for comparison."}
        newer_pages = json.loads(rows[0]["results_json"] or "[]")
        older_pages = json.loads(rows[1]["results_json"] or "[]")
    except Exception as exc:
        return {"has_diff": False, "message": f"Could not load snapshot data: {exc}"}

    # Build issue maps: {url: set(issues)}
    def issue_map(pages):
        m = {}
        for p in pages:
            iss = p.get("issues", [])
            if iss:
                m[p["url"]] = set(iss)
        return m

    old_map = issue_map(older_pages)
    new_map = issue_map(newer_pages)

    # New issues: appeared in newer but not older (or new URL with issues)
    new_issues = []
    for url, issues in new_map.items():
        old_iss = old_map.get(url, set())
        appeared = issues - old_iss
        for iss in appeared:
            new_issues.append({"url": url, "issue": iss, "type": "new"})

    # Fixed issues: were in older but not newer
    fixed_issues = []
    for url, issues in old_map.items():
        new_iss = new_map.get(url, set())
        resolved = issues - new_iss
        for iss in resolved:
            fixed_issues.append({"url": url, "issue": iss, "type": "fixed"})

    score_delta = round(
        (newer.get("health_score") or 0) - (older.get("health_score") or 0), 1
    )

    return {
        "has_diff":    True,
        "newer_date":  newer.get("crawled_at", ""),
        "older_date":  older.get("crawled_at", ""),
        "score_delta": score_delta,
        "new_issues":  new_issues[:100],
        "fixed_issues": fixed_issues[:100],
        "new_issue_count":   len(new_issues),
        "fixed_issue_count": len(fixed_issues),
    }


# ── Team / workspace endpoints ────────────────────────────────────────────────

class TeamInviteRequest(BaseModel):
    project_id: int
    email:      str = Field(..., min_length=5)
    role:       str = Field(default="viewer", pattern="^(viewer|editor)$")

class TeamRemoveRequest(BaseModel):
    project_id: int
    email:      str

@app.post("/team/invite")
def team_invite(req: TeamInviteRequest, request: _FastAPIRequest):
    """
    Share a project with another user by email.
    The invited user must have a CrawlIQ account.
    Role: viewer (read-only) or editor (can save snapshots).
    """
    if not _AUTH_MODULE or not _PROJECTS_MODULE:
        raise HTTPException(503, "Auth or Projects module not available")
    owner = _get_current_user(request)
    if not owner:
        raise HTTPException(401, "Not authenticated")

    proj = _db_get_project(req.project_id)
    if not proj:
        raise HTTPException(404, "Project not found")
    if proj.get("user_id") != owner["id"]:
        raise HTTPException(403, "Only the project owner can invite members")

    # Find the invited user
    try:
        from competitor_db import _connect as _db_conn
        with _db_conn() as conn:
            invitee = conn.execute(
                "SELECT id, name, email FROM users WHERE email=?", (req.email.lower(),)
            ).fetchone()
        if not invitee:
            raise HTTPException(404, f"No account found for {req.email}. They need to register first.")
        invitee = dict(invitee)
        # Upsert team membership
        from datetime import datetime as _dt, timezone as _tz
        with _db_conn() as conn:
            conn.execute("""
                INSERT INTO project_members (project_id, user_id, role, invited_at)
                VALUES (?,?,?,?)
                ON CONFLICT(project_id, user_id) DO UPDATE SET role=excluded.role
            """, (req.project_id, invitee["id"], req.role, _dt.now(_tz.utc).isoformat(timespec="seconds")))
        return {
            "status": "invited",
            "user": {"name": invitee["name"], "email": invitee["email"]},
            "role": req.role,
            "project_id": req.project_id,
        }
    except HTTPException:
        raise
    except Exception as exc:
        # project_members table may not exist yet — handle gracefully
        if "no such table" in str(exc).lower():
            raise HTTPException(503, "Team feature requires DB migration. Restart the server.")
        raise HTTPException(500, str(exc))


@app.get("/team/members/{project_id}")
def team_list_members(project_id: int, request: _FastAPIRequest):
    """List all members with access to a project. Auth required."""
    if not _PROJECTS_MODULE:
        raise HTTPException(503, "Projects module not available")
    if _AUTH_MODULE:
        user = _get_current_user(request)
        if not user:
            raise HTTPException(401, "Not authenticated")
        proj = _db_get_project(project_id)
        if proj and proj.get("user_id") and proj["user_id"] != user["id"]:
            try:
                from competitor_db import _connect as _dbc
                with _dbc() as conn:
                    member = conn.execute(
                        "SELECT 1 FROM project_members WHERE project_id=? AND user_id=?",
                        (project_id, user["id"])
                    ).fetchone()
                if not member:
                    raise HTTPException(403, "Access denied")
            except HTTPException:
                raise
            except Exception:
                raise HTTPException(403, "Access denied")
    try:
        from competitor_db import _connect as _db_conn
        with _db_conn() as conn:
            rows = conn.execute("""
                SELECT u.id, u.name, u.email, pm.role, pm.invited_at
                FROM project_members pm
                JOIN users u ON u.id = pm.user_id
                WHERE pm.project_id=?
            """, (project_id,)).fetchall()
        return {"members": [dict(r) for r in rows]}
    except Exception as exc:
        if "no such table" in str(exc).lower():
            return {"members": []}
        raise HTTPException(500, str(exc))


@app.delete("/team/member")
def team_remove_member(req: TeamRemoveRequest, request: _FastAPIRequest):
    """Remove a user's access to a project."""
    if not _AUTH_MODULE:
        raise HTTPException(503, "Auth module not available")
    owner = _get_current_user(request)
    if not owner:
        raise HTTPException(401, "Not authenticated")
    try:
        from competitor_db import _connect as _db_conn
        with _db_conn() as conn:
            invitee = conn.execute("SELECT id FROM users WHERE email=?", (req.email.lower(),)).fetchone()
            if not invitee:
                raise HTTPException(404, "User not found")
            conn.execute(
                "DELETE FROM project_members WHERE project_id=? AND user_id=?",
                (req.project_id, invitee["id"])
            )
        return {"status": "removed"}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, str(exc))


def _handle_sigterm(sig, frame):
    """
    BUG-N19: graceful SIGTERM — mark crawl as stopped so the next startup
    does not show a stale running=True flag from the previous container.
    """
    logger.info("SIGTERM received — shutting down gracefully.")
    crawl_status.update({"running": False, "done": False, "error": "Server shutdown"})
    sys.exit(0)

signal.signal(signal.SIGTERM, _handle_sigterm)


if __name__ == "__main__":
    parser = _build_arg_parser()
    args   = parser.parse_args()

    if args.crawl:
        # ── CLI PIPELINE MODE ────────────────────────────────────────────
        _run_cli(args)
    else:
        # ── WEB SERVER MODE (default) ────────────────────────────────────
        provider = os.getenv("AI_PROVIDER", "groq").upper()
        key_set  = _ai_configured()
        print()
        print("╔══════════════════════════════════════════════════════════╗")
        print("║       SEO Crawler Dashboard — FastAPI Server             ║")
        print("╠══════════════════════════════════════════════════════════╣")
        print(f"║  URL:      http://{args.host}:{args.port}{'':24s}   ║")
        print(f"║  AI:       {provider} ({'key set ✓' if key_set else 'NO KEY — set env var'}){'':20s}   ║")
        print("║  HTML UI:  http://localhost:" + str(args.port) + "          (index.html)   ║")
        print("╚══════════════════════════════════════════════════════════╝")
        print()
        uvicorn.run(
            "main:app",
            host=args.host,
            port=args.port,
            reload=args.reload,
        )
